#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Bot de envío automático de reportes de rendimiento por cliente.

Flujo:
  1) Lee la base de clientes (Excel) con columnas CLIENTE y CORREO.
  2) Por cada fila, genera el reporte usando la plantilla .xlsm y la macro VBA
     (Excel instalado en Windows, mediante COM). La salida se guarda como .xlsx.
  3) Envía un correo por cliente con el adjunto correspondiente.

Ejecución one-shot: sin bucles de programación interna (Cron / Programador de tareas).

Variables de entorno importantes (generación con macro):
  EXCEL_TEMPLATE_XLSM   Ruta absoluta a FRF_Rend ACTUALIZADO.xlsm (plantilla con macros).
  REPORT_MACRO_RUN      Nombre para Application.Run, p.ej. 'GenerarReporte' o
                        "'FRF_Rend ACTUALIZADO.xlsm'!GenerarReporte".
  REPORT_CLIENT_CELL    Celda donde escribir el nombre del cliente antes de la macro,
                        p.ej. "Parametros!B2" (hoja y celda según tu libro).
  REPORT_OUTPUT_DIR     Carpeta donde guardar cada .xlsx generado (por defecto carpeta temporal).

  CLIENTES_XLSX_PATH    Ruta a "BASES DE DATOS CLIENTES COLBEEF.xlsx".
  SMTP_*                Credenciales de correo (ver función obtener_config_smtp).

Opcionales:
  MAX_ENVIOS            Número máximo de correos a enviar en una corrida (pruebas).
  DRY_RUN               Si es "1", genera (si aplica) pero no envía correos.
  SKIP_GENERATION       Si es "1", no llama a Excel; útil para probar solo lectura de clientes.

  Por defecto el envío va a TODOS los correos de la celda CORREO (p. ej. cliente + servicioalcliente).

  EJECUCION_NOCTURNA   Si es "1", tras cargar .env fuerza envío real (DRY_RUN=0), quita límite MAX_ENVIOS,
                       limpia DB_PROPIETARIO_LIKE (cada cliente usa su nombre del Excel en la consulta) y,
                       si no hay DB_FECHA_PLAN_AUTO, usa "ayer" (faena del día anterior a las 03:00).

  DB_FECHA_PLAN_AUTO   hoy | ayer | today | yesterday — ignora DB_FECHA_PLAN para {{FECHA_PLAN}}.

  SOLO_CORREO_EXTERNO_CLIENTE  Si es "1", se excluyen dominios internos (ver abajo). Por defecto desactivado.

  EMAIL_EXCLUDED_DOMAINS   Solo si SOLO_CORREO_EXTERNO_CLIENTE=1. Dominios a excluir (coma). Por defecto: colbeef.com
  CLIENTE_FINAL_UN_SOLO    Solo con filtro externo: si es "1", un solo correo externo (el primero de la lista).

  CLIENTES_LISTA_ORIGEN   excel (defecto) | sirt | merge. Solo con USE_DB_QUERY=1 y POSTGRES_*.
                          sirt: propietarios activos en FECHA_PLAN con correos desde organizaciones.empresa
                          (cambia de tercero en SIRT sin depender del Excel).
                          merge: misma lista que sirt y además añade correos de filas del Excel con nombre compatible.

  CLIENTES_SOLO_CON_CAVA  Con lista sirt|merge: por defecto "1" (solo clientes con al menos una cava
                          en la faena). Pon "0" para incluir todo propietario con sacrificio aunque no
                          tenga fila de cava. Con lista excel no aplica aquí.
                          Criterio de cava: nombre_cava no vacío, o destino con texto cava, o destino
                          tipo numerado con prefijo día (p. ej. "02085 /LxM/").

  CLIENTES_LOG_CAVAS    Si es "1", en cada cliente (modo BD) registra en log las cavas detectadas.

  SMTP_BCC              Opcional. Lista de correos en copia oculta (coma o ;) en cada envío, p. ej. equipo interno
                        para monitorear un piloto sin redirigir el Para del cliente.

  EMAIL_SUBJECT_PREFIX  Prefijo del asunto; si incluye "prueba" o "piloto", el cuerpo HTML puede mostrar aviso piloto.
  EMAIL_AVISO_PILOTO_CUERPO  Si es "1" (defecto) y el prefijo contiene prueba/piloto, se inserta un recuadro de aviso
                        en el correo. Pon "0" para ocultar ese bloque.

  REPORT_SUMMARY_EMAIL  Tras procesar todos los clientes, envía un correo HTML de resumen (sin adjunto) a esta
                        lista (coma o ;), dominio @colbeef.com. Por defecto: desarrollo.tecnologia y coordinacion.linea.
                        Pon "0" o "false" para no enviar resumen.

  GOOGLE_DRIVE_REPORTES_PARENT  Ruta local de "Mi unidad" / Google Drive (Drive para escritorio).
                                Si está definida, tras generar cada adjunto se copia a:
                                {parent}/Reportes/{YYYY-MM-DD}/archivo
                                (fecha = beneficio DB_FECHA_PLAN / auto, o hoy si no hay).
                                No usa API ni correo: solo sincronización local de Google.
  DRIVE_REPORTES_FOLDER_NAME    Opcional; por defecto Reportes (subcarpeta bajo parent).
"""

from __future__ import annotations

import html
import logging
import os
import re
import smtplib
import sys
import subprocess
import shutil
import unicodedata
from email.utils import formataddr
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any, List, Optional, Sequence, Set, Tuple

# openpyxl: lectura de la base de clientes (y documentado para plantillas con keep_vba)
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# Copia en ciego opcional (administradores que reciben todos los envíos)
# Dejar vacío si no aplica.
# -----------------------------------------------------------------------------
COPIA_OCULTA: List[str] = []

# Ruta por defecto a la base de clientes (sobrescribible con CLIENTES_XLSX_PATH)
_DEFAULT_CLIENTES = Path.home() / "Downloads" / "BASES DE DATOS CLIENTES COLBEEF.xlsx"

# Plantilla con macros (sobrescribible con EXCEL_TEMPLATE_XLSM)
_DEFAULT_TEMPLATE = Path.home() / "Downloads" / "FRF_Rend ACTUALIZADO.xlsm"


@dataclass
class ClienteFila:
    """Un registro de cliente leído de la base."""

    nombre: str
    correos: List[str]
    fila_excel: int
    origen: str = "excel"  # excel | sirt | merge
    cavas_resumen: str = ""  # etiquetas agregadas desde SIRT (solo listas sirt|merge cuando aplica)


def configurar_logging(ruta_log: Path) -> None:
    """Configura logging en archivo y consola con hora detallada."""
    ruta_log.parent.mkdir(parents=True, exist_ok=True)
    formato = "%(asctime)s | %(levelname)s | %(message)s"
    fecha_fmt = "%Y-%m-%d %H:%M:%S"

    root = logging.getLogger()
    root.setLevel(logging.DEBUG)

    manejador_archivo = logging.FileHandler(ruta_log, encoding="utf-8")
    manejador_archivo.setLevel(logging.DEBUG)
    manejador_archivo.setFormatter(logging.Formatter(formato, fecha_fmt))

    manejador_consola = logging.StreamHandler(sys.stdout)
    manejador_consola.setLevel(logging.INFO)
    manejador_consola.setFormatter(logging.Formatter(formato, fecha_fmt))

    root.handlers.clear()
    root.addHandler(manejador_archivo)
    root.addHandler(manejador_consola)


def directorio_aplicacion() -> Path:
    """
    Carpeta donde están el .env, el log y ultima_corrida.txt.
    Con PyInstaller (--onefile/--onedir) apunta junto al .exe; en desarrollo, junto a este .py.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def cargar_dotenv_proyecto() -> None:
    """Lee `.env` junto a este script y aplica variables (prioridad sobre entorno vacío)."""
    env_path = directorio_aplicacion() / ".env"
    if not env_path.is_file():
        return
    try:
        texto = env_path.read_text(encoding="utf-8")
    except OSError:
        return
    for line in texto.splitlines():
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        if "=" not in s:
            continue
        clave, _, valor = s.partition("=")
        clave = clave.strip()
        valor = valor.strip().strip('"').strip("'")
        if clave:
            os.environ[clave] = valor


def obtener_ruta_clientes() -> Path:
    """Resuelve la ruta al Excel de base de clientes."""
    default = _DEFAULT_CLIENTES.resolve()
    env = os.getenv("CLIENTES_XLSX_PATH", "").strip()
    if env:
        p = Path(env).expanduser().resolve()
        if p.is_file():
            return p
        # Placeholder típico o archivo en otra carpeta: intentar Descargas
        env_lower = str(p).lower()
        if "ruta_real" in env_lower or "c:\\ruta" in env_lower:
            if default.is_file():
                return default
        if not p.exists() and default.is_file():
            return default
        return p
    if default.is_file():
        return default
    return default


def obtener_ruta_plantilla_xlsm() -> Path:
    """Resuelve la ruta a la plantilla .xlsm con macros."""
    env = os.getenv("EXCEL_TEMPLATE_XLSM")
    if env:
        return Path(env).expanduser().resolve()
    return _DEFAULT_TEMPLATE.resolve()


def obtener_directorio_salida() -> Path:
    """Directorio donde guardar los reportes .xlsx generados."""
    env = os.getenv("REPORT_OUTPUT_DIR")
    if env:
        p = Path(env).expanduser().resolve()
    else:
        p = Path(tempfile.gettempdir()) / "rendimientos_bot"
    p.mkdir(parents=True, exist_ok=True)
    return p


def respaldar_reporte_en_unidad_sincronizada(
    archivo: Path,
    *,
    fecha_subcarpeta: str,
    log: logging.Logger,
) -> None:
    """
    Copia el reporte generado a una ruta bajo la carpeta local de Google Drive
    (Drive para escritorio / File Stream), para que suba sola a la nube.

    Estructura: {GOOGLE_DRIVE_REPORTES_PARENT}/{DRIVE_REPORTES_FOLDER_NAME}/{fecha}/archivo

    No requiere credenciales OAuth ni cuenta SMTP aparte: es una copia de archivos
    a un directorio que Google ya sincroniza.
    """
    raw = os.getenv("GOOGLE_DRIVE_REPORTES_PARENT", "").strip().strip('"').strip("'")
    if not raw:
        return
    if not archivo.is_file():
        log.warning("Respaldo Drive omitido: no existe el archivo %s", archivo)
        return
    try:
        parent = Path(raw).expanduser().resolve()
    except OSError as e:
        log.warning("GOOGLE_DRIVE_REPORTES_PARENT inválida (%s): %s", raw, e)
        return
    sub = (os.getenv("DRIVE_REPORTES_FOLDER_NAME", "Reportes").strip() or "Reportes").strip()
    # Fecha YYYY-MM-DD (solo nombre de carpeta)
    fd = (fecha_subcarpeta or "")[:10].replace("/", "-").strip() or date.today().strftime("%Y-%m-%d")
    dest_dir = parent / sub / fd
    try:
        dest_dir.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        log.warning("No se pudo crear carpeta de respaldo Drive %s: %s", dest_dir, e)
        return
    destino = dest_dir / archivo.name
    try:
        if destino.resolve() == archivo.resolve():
            return
    except OSError:
        pass
    try:
        shutil.copy2(archivo, destino)
        log.info("Respaldo en unidad sincronizada: %s", destino)
    except OSError as e:
        log.warning("No se pudo copiar respaldo a Drive local %s: %s", destino, e)


def _normalizar_encabezado(val) -> str:
    if val is None:
        return ""
    return str(val).strip().upper()


def detectar_columnas_cliente_correo(hoja) -> Tuple[int, int]:
    """
    Encuentra índices de columna (0-based) para CLIENTE y CORREO en las primeras filas.
    """
    for fila in hoja.iter_rows(min_row=1, max_row=10, values_only=True):
        if not fila:
            continue
        celdas = list(fila)
        mapa = {_normalizar_encabezado(c): i for i, c in enumerate(celdas) if c is not None}
        if "CLIENTE" in mapa and "CORREO" in mapa:
            return mapa["CLIENTE"], mapa["CORREO"]
    raise ValueError(
        "No se encontraron las columnas CLIENTE y CORREO en las primeras filas de la hoja."
    )


def parsear_lista_correos(texto: Optional[str]) -> List[str]:
    """
    Separa correos separados por ; o , y elimina duplicados conservando orden.
    """
    if not texto or not str(texto).strip():
        return []
    trozos = re.split(r"[;,]", str(texto))
    resultado: List[str] = []
    vistos = set()
    for t in trozos:
        correo = t.strip()
        if not correo or "@" not in correo:
            continue
        clave = correo.lower()
        if clave not in vistos:
            vistos.add(clave)
            resultado.append(correo)
    return resultado


def _parse_env_emails(var_name: str) -> List[str]:
    """Lee una variable env (csv/; separado) y devuelve lista de correos."""
    raw = os.getenv(var_name, "").strip()
    return parsear_lista_correos(raw)


def solo_correo_externo_cliente_activo() -> bool:
    """Si True, no se envía a dominios internos listados en EMAIL_EXCLUDED_DOMAINS."""
    return os.getenv("SOLO_CORREO_EXTERNO_CLIENTE", "").strip().lower() in (
        "1",
        "true",
        "yes",
    )


def obtener_dominios_correo_interno_excluidos() -> Set[str]:
    """
    Dominios considerados internos cuando SOLO_CORREO_EXTERNO_CLIENTE=1.
    EMAIL_EXCLUDED_DOMAINS: lista separada por comas, p.ej. colbeef.com
    """
    raw = os.getenv("EMAIL_EXCLUDED_DOMAINS", "colbeef.com")
    return {x.strip().lower() for x in raw.split(",") if x.strip()}


def correo_es_interno(correo: str, dominios_excluidos: Set[str]) -> bool:
    """True si el dominio del correo coincide con alguno excluido (subdominios incluidos)."""
    if "@" not in correo:
        return True
    dom = correo.split("@", 1)[1].lower().strip()
    for d in dominios_excluidos:
        if dom == d or dom.endswith("." + d):
            return True
    return False


def filtrar_correos_cliente_final(
    correos: List[str],
    log: Optional[logging.Logger],
    dominios_excluidos: Optional[Set[str]] = None,
) -> List[str]:
    """
    Deja solo correos del cliente final (excluye dominios internos tipo Colbeef).
    """
    doms = dominios_excluidos if dominios_excluidos is not None else obtener_dominios_correo_interno_excluidos()
    externos = [c for c in correos if not correo_es_interno(c, doms)]
    if (
        log
        and len(correos) != len(externos)
        and len(correos) > 0
    ):
        log.debug(
            "Filtrado cliente final: %s -> %s (excluidos dominios: %s)",
            correos,
            externos,
            sorted(doms),
        )
    solo_uno = os.getenv("CLIENTE_FINAL_UN_SOLO", "").strip().lower() in (
        "1",
        "true",
        "yes",
    )
    if solo_uno and externos:
        return [externos[0]]
    return externos


def cargar_clientes_desde_excel(ruta: Path, log: logging.Logger) -> List[ClienteFila]:
    """
    Lee la hoja TablaContactos (o la primera hoja si no existe el nombre)
    y devuelve una lista de ClienteFila.
    """
    if not ruta.is_file():
        raise FileNotFoundError(f"No se encontró la base de clientes: {ruta}")

    # read_only=False: permite detectar cabeceras y recorrer filas sin conflictos de iterador.
    wb = openpyxl.load_workbook(ruta, read_only=False, data_only=True)
    try:
        if "TablaContactos" in wb.sheetnames:
            hoja = wb["TablaContactos"]
        else:
            log.warning(
                "No existe la hoja 'TablaContactos'; se usa la primera hoja: %s",
                wb.sheetnames[0],
            )
            hoja = wb[wb.sheetnames[0]]

        idx_cliente, idx_correo = detectar_columnas_cliente_correo(hoja)
        log.info(
            "Columnas detectadas: CLIENTE (índice %s), CORREO (índice %s)",
            idx_cliente,
            idx_correo,
        )
        filtrar_internos = solo_correo_externo_cliente_activo()
        dominios_excl = (
            obtener_dominios_correo_interno_excluidos() if filtrar_internos else frozenset()
        )
        if filtrar_internos:
            log.info(
                "Modo solo correo externo: dominios excluidos del envío: %s",
                sorted(dominios_excl),
            )
        else:
            log.info(
                "Envío a todos los correos de la base en cada fila (p. ej. cliente e internos Colbeef)."
            )

        clientes: List[ClienteFila] = []
        for i, fila in enumerate(
            hoja.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not fila:
                continue
            celdas = list(fila)
            if idx_cliente >= len(celdas):
                continue
            nombre = celdas[idx_cliente]
            correo_raw = celdas[idx_correo] if idx_correo < len(celdas) else None
            if nombre is None or not str(nombre).strip():
                continue
            nombre_limpio = str(nombre).strip()
            correos_raw = parsear_lista_correos(correo_raw)
            if filtrar_internos:
                correos = filtrar_correos_cliente_final(
                    correos_raw, log, set(dominios_excl)
                )
            else:
                correos = list(correos_raw)
            if not correos:
                if correos_raw and filtrar_internos:
                    log.warning(
                        "Fila %s: cliente '%s' solo tiene correos internos (%s); se omite.",
                        i,
                        nombre_limpio[:80],
                        correos_raw,
                    )
                elif not correos_raw:
                    log.warning(
                        "Fila %s: cliente '%s' sin correos válidos; se omite.",
                        i,
                        nombre_limpio[:80],
                    )
                continue
            clientes.append(
                ClienteFila(
                    nombre=nombre_limpio,
                    correos=correos,
                    fila_excel=i,
                    origen="excel",
                )
            )
    finally:
        wb.close()

    log.info("Clientes cargados con correo válido: %s", len(clientes))
    return clientes


def _recopilar_correos_unicos(correos: Sequence[str]) -> List[str]:
    """Orden estable, sin duplicados (comparación por minúsculas)."""
    vistos: Set[str] = set()
    out: List[str] = []
    for c in correos:
        cl = (c or "").strip()
        if not cl or "@" not in cl:
            continue
        k = cl.lower()
        if k not in vistos:
            vistos.add(k)
            out.append(cl)
    return out


def _nombres_cliente_coinciden_fuzzy(nombre_a: str, nombre_b: str) -> bool:
    """Coincidencia flexible entre nombre Excel y nombre empresa SIRT."""
    na = _normalizar_nombre(nombre_a)
    nb = _normalizar_nombre(nombre_b)
    if not na or not nb:
        return False
    return na == nb or na in nb or nb in na


def combinar_correos_cliente_sirt_con_excel(
    clientes_sirt: List[ClienteFila],
    clientes_excel: List[ClienteFila],
    log: logging.Logger,
) -> List[ClienteFila]:
    """
    Lista de envío = SIRT (nombre y correos de empresa); añade correos del Excel
    cuando el nombre de la fila coincide de forma flexible con el de SIRT.
    """
    resultado: List[ClienteFila] = []
    for sc in clientes_sirt:
        merged = list(sc.correos)
        for ec in clientes_excel:
            if not _nombres_cliente_coinciden_fuzzy(sc.nombre, ec.nombre):
                continue
            antes = len(merged)
            merged = _recopilar_correos_unicos(merged + ec.correos)
            if len(merged) > antes:
                log.debug(
                    "Merge Excel→SIRT: propietario SIRT '%s' incorporó correos desde Excel fila %s ('%s')",
                    sc.nombre[:80],
                    ec.fila_excel,
                    ec.nombre[:80],
                )
        resultado.append(
            ClienteFila(
                nombre=sc.nombre,
                correos=merged,
                fila_excel=sc.fila_excel,
                origen="merge",
                cavas_resumen=sc.cavas_resumen,
            )
        )
    log.info(
        "Lista merge: %s clientes desde SIRT con posible ampliación de correos desde Excel.",
        len(resultado),
    )
    return resultado


def _resolver_solo_cava_lista_sirt() -> bool:
    """
    True = solo empresas con al menos una fila de cava en vw_pbi01 para la faena.
    Por defecto activo para listas desde BD; desactivar con CLIENTES_SOLO_CON_CAVA=0|false.
    """
    raw = os.getenv("CLIENTES_SOLO_CON_CAVA", "").strip().lower()
    if raw in ("0", "false", "no", "off"):
        return False
    if raw in ("1", "true", "yes", "on"):
        return True
    lo = os.getenv("CLIENTES_LISTA_ORIGEN", "excel").strip().lower()
    return lo in ("sirt", "merge")


def cargar_clientes_desde_sirt(log: logging.Logger) -> List[ClienteFila]:
    """
    Propietarios con faena en FECHA_PLAN y sacrificio en [fecha, fecha+2d),
    con correos desde organizaciones.empresa (producto_empresa activo).

    Alineado al prefiltro por vw_pbi01 para que los nombres coincidan con la consulta de rendimientos.
    Por defecto (lista sirt|merge) solo incluye clientes con al menos una asignación de cava
    (nombre_cava, destino con 'cava', o destino tipo '02085 /LxM/').
    """
    fecha_plan = resolver_db_fecha_plan()
    if not fecha_plan:
        raise ValueError(
            "Para CLIENTES_LISTA_ORIGEN=sirt|merge define DB_FECHA_PLAN o DB_FECHA_PLAN_AUTO (hoy/ayer)."
        )

    solo_cava = _resolver_solo_cava_lista_sirt()
    filtrar_internos = solo_correo_externo_cliente_activo()
    dominios_excl = (
        obtener_dominios_correo_interno_excluidos() if filtrar_internos else frozenset()
    )

    sql = """
        WITH base AS (
            SELECT
                e.id AS empresa_id,
                e.nombre AS empresa_nombre,
                e.correo,
                e.correos_opcionales,
                e.correo_facturacion,
                COALESCE(
                    NULLIF(BTRIM(v.nombre_cava::text), ''),
                    NULLIF(BTRIM(v.destino::text), '')
                ) AS etiqueta,
                (
                    NULLIF(BTRIM(v.nombre_cava::text), '') IS NOT NULL
                    OR COALESCE(v.destino::text, '') ILIKE '%%cava%%'
                    OR COALESCE(BTRIM(v.destino::text), '')
                        ~ '/[[:alnum:]]{1,4}[xX][[:alnum:]]{1,4}/'
                ) AS es_fila_cava
            FROM trazabilidad_proceso.plan_faena pf
            JOIN trazabilidad_proceso.plan_faena_producto pfp
              ON pfp.id_plan_faena = pf.id
            JOIN trazabilidad_proceso.vw_pbi01 v
              ON v.codigo::text = pfp.id_producto::text
            JOIN trazabilidad_proceso.producto_empresa pe
              ON pe.id_producto::text = pfp.id_producto::text
             AND pe.activo = true
            JOIN organizaciones.empresa e
              ON e.id = pe.id_empresa
            WHERE pf.fecha_plan = %s::date
              AND v.fecha_insensibilizacion >= %s::timestamp
              AND v.fecha_insensibilizacion < (%s::date + INTERVAL '2 days')
              AND NULLIF(BTRIM(e.nombre), '') IS NOT NULL
        ),
        labels AS (
            SELECT DISTINCT empresa_id, etiqueta
            FROM base
            WHERE es_fila_cava
              AND NULLIF(BTRIM(etiqueta), '') IS NOT NULL
        ),
        emp AS (
            SELECT DISTINCT ON (empresa_id)
                empresa_id,
                empresa_nombre,
                correo,
                correos_opcionales,
                correo_facturacion
            FROM base
            ORDER BY empresa_id, empresa_nombre
        ),
        por_empresa AS (
            SELECT
                e.empresa_id,
                e.empresa_nombre,
                e.correo,
                e.correos_opcionales,
                e.correo_facturacion,
                STRING_AGG(l.etiqueta, ' | ' ORDER BY l.etiqueta) AS cavas_resumen,
                COUNT(l.etiqueta) > 0 AS tiene_alguna_cava
            FROM emp e
            LEFT JOIN labels l ON l.empresa_id = e.empresa_id
            GROUP BY e.empresa_id, e.empresa_nombre, e.correo, e.correos_opcionales, e.correo_facturacion
        )
        SELECT
            empresa_id,
            empresa_nombre,
            correo,
            correos_opcionales,
            correo_facturacion,
            cavas_resumen
        FROM por_empresa
        WHERE (NOT %s::boolean) OR tiene_alguna_cava
        ORDER BY empresa_nombre
    """

    psycopg = _obtener_psycopg()
    dsn_log, kwargs = _conn_postgres_desde_env()
    clientes: List[ClienteFila] = []
    log.debug("Clientes SIRT: conectando a PostgreSQL: %s", dsn_log)
    with psycopg.connect(**kwargs) as conn:
        with conn.cursor() as cur:
            cur.execute(
                sql,
                (fecha_plan, fecha_plan, fecha_plan, solo_cava),
            )
            rows = cur.fetchall()
    for row in rows:
        _eid, nombre, correo, correos_opc, correo_fact, cavas_txt = row
        nombre_limpio = str(nombre).strip() if nombre else ""
        if not nombre_limpio:
            continue
        correos_raw: List[str] = []
        for campo in (correo, correos_opc, correo_fact):
            correos_raw.extend(parsear_lista_correos(campo))
        correos_raw = _recopilar_correos_unicos(correos_raw)
        if filtrar_internos:
            correos = filtrar_correos_cliente_final(
                correos_raw, log, set(dominios_excl)
            )
        else:
            correos = list(correos_raw)
        if not correos:
            if correos_raw and filtrar_internos:
                log.warning(
                    "SIRT cliente '%s' solo tiene correos internos (%s); se omite.",
                    nombre_limpio[:80],
                    correos_raw,
                )
            elif not correos_raw:
                log.warning(
                    "SIRT cliente '%s' sin correos en empresa (correo/correos_opcionales); se omite.",
                    nombre_limpio[:80],
                )
            continue
        cavas_resumen = (str(cavas_txt).strip() if cavas_txt is not None else "") or ""
        clientes.append(
            ClienteFila(
                nombre=nombre_limpio,
                correos=correos,
                fila_excel=0,
                origen="sirt",
                cavas_resumen=cavas_resumen,
            )
        )

    con_cava = sum(1 for c in clientes if c.cavas_resumen)
    log.info(
        "Clientes desde SIRT (fecha_plan=%s, solo_cava=%s): total=%s, con resumen de cavas=%s",
        fecha_plan,
        solo_cava,
        len(clientes),
        con_cava,
    )
    mostrados = 0
    for muestra in clientes:
        if not muestra.cavas_resumen:
            continue
        log.info(
            "Ejemplo cavas — %s: %s",
            muestra.nombre[:80],
            muestra.cavas_resumen[:400],
        )
        mostrados += 1
        if mostrados >= 5:
            break
    return clientes


def _normalizar_nombre_cliente(s: str) -> str:
    s = (s or "").strip().lower()
    # Normalización mínima para matching robusto
    return re.sub(r"\s+", " ", s)


def nombre_archivo_reporte(nombre_cliente: str, fecha: date) -> str:
    """Nombre del adjunto alineado al ejemplo: Rendimientos CLIENTE Beneficio YYYY-MM-DD.<ext>"""
    f = fecha.strftime("%Y-%m-%d")
    ext = os.getenv("REPORT_FILE_EXT", "xlsx").strip().lower().lstrip(".") or "xlsx"
    if ext not in ("xlsx", "xlsm", "pdf"):
        ext = "xlsx"
    return f"Rendimientos {nombre_cliente} Beneficio {f}.{ext}"


def sanitizar_nombre_archivo(nombre: str) -> str:
    """Evita caracteres no permitidos en rutas Windows."""
    return re.sub(r'[<>:"/\\|?*]', "_", nombre)


def generar_reporte_con_excel_com(
    plantilla_xlsm: Path,
    nombre_cliente: str,
    ruta_salida: Path,
    log: logging.Logger,
) -> None:
    """
    Abre la plantilla .xlsm con Excel, escribe el cliente en REPORT_CLIENT_CELL,
    ejecuta REPORT_MACRO_RUN y guarda el resultado como .xlsx en ruta_salida.

    Requiere Windows, Microsoft Excel y pywin32 instalados.
    """
    try:
        import win32com.client  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "Para ejecutar macros VBA se requiere 'pywin32' y Excel en Windows. "
            "Instala con: pip install pywin32"
        ) from e

    macro_run = os.getenv("REPORT_MACRO_RUN", "").strip()
    celda_cliente = os.getenv("REPORT_CLIENT_CELL", "").strip()
    if not macro_run:
        raise ValueError(
            "Define REPORT_MACRO_RUN con el nombre de la macro (Application.Run)."
        )
    if not celda_cliente:
        raise ValueError(
            'Define REPORT_CLIENT_CELL, por ejemplo "Parametros!B2" según tu plantilla.'
        )

    if not plantilla_xlsm.is_file():
        raise FileNotFoundError(f"Plantilla no encontrada: {plantilla_xlsm}")

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    if ruta_salida.is_file():
        ruta_salida.unlink()

    xl_path_template = str(plantilla_xlsm.resolve())
    xl_path_out = str(ruta_salida.resolve())

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(xl_path_template)
        if "!" in celda_cliente:
            hoja_nombre, rango = celda_cliente.split("!", 1)
            hoja_nombre = hoja_nombre.strip().strip("'\"")
            rango = rango.strip()
            ws = wb.Worksheets(hoja_nombre)
            ws.Range(rango).Value = nombre_cliente
        else:
            raise ValueError(
                'REPORT_CLIENT_CELL debe incluir hoja y celda, p.ej. "Datos!B2"'
            )

        log.debug("Ejecutando macro: %s", macro_run)
        excel.Application.Run(macro_run)

        # 51 = xlOpenXMLWorkbook (.xlsx), sin macros en el adjunto al cliente
        wb.SaveAs(xl_path_out, FileFormat=51)
        log.info("Reporte guardado: %s", ruta_salida)
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        excel.Quit()


def generar_excel_demo(ruta_salida: Path, log: logging.Logger) -> None:
    """
    Genera un .xlsx mínimo para validar que el flujo de escritura funciona.
    No requiere plantilla .xlsm, macro, ni base de clientes.
    """
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "DEMO"
    ws["A1"] = "OK"
    ws["B1"] = "Archivo generado por bot_reportes.py"
    ws["A2"] = "Fecha"
    ws["B2"] = date.today().strftime("%Y-%m-%d")
    wb.save(ruta_salida)
    log.info("Excel DEMO guardado: %s", ruta_salida)


def _leer_sql_desde_env_o_archivo() -> str:
    """
    Lee SQL desde DB_QUERY o desde DB_QUERY_FILE.
    DB_QUERY_FILE puede apuntar a un .sql con el query completo.
    """
    p = os.getenv("DB_QUERY_FILE", "").strip()
    if p:
        return Path(p).expanduser().resolve().read_text(encoding="utf-8")
    q = os.getenv("DB_QUERY", "").strip()
    if q:
        return q
    raise ValueError("Define DB_QUERY o DB_QUERY_FILE para ejecutar la consulta SQL.")


def resolver_db_fecha_plan() -> Optional[str]:
    """
    Resuelve la fecha para el placeholder {{FECHA_PLAN}}.
    Prioridad: DB_FECHA_PLAN_AUTO (hoy/ayer) y si no, DB_FECHA_PLAN fija.
    """
    auto = os.getenv("DB_FECHA_PLAN_AUTO", "").strip().lower()
    if auto in ("hoy", "today"):
        return date.today().strftime("%Y-%m-%d")
    if auto in ("ayer", "yesterday"):
        return (date.today() - timedelta(days=1)).strftime("%Y-%m-%d")
    fija = os.getenv("DB_FECHA_PLAN", "").strip()
    return fija or None


def aplicar_modo_ejecucion_nocturna(log: logging.Logger) -> None:
    """
    Ejecución programada (p. ej. 03:00): todos los clientes, envío real,
    propietario = nombre del Excel por fila (DB_PROPIETARIO_LIKE vacío).
    """
    if os.getenv("EJECUCION_NOCTURNA", "").strip().lower() not in ("1", "true", "yes"):
        return
    os.environ.pop("DEMO_XLSX", None)
    os.environ["DRY_RUN"] = "0"
    os.environ.pop("MAX_ENVIOS", None)
    os.environ.pop("SOLO_CLIENTE_NOMBRE", None)
    os.environ["DB_PROPIETARIO_LIKE"] = ""
    # Si TEST_EMAIL_TO viene informado, se usa como redireccion global controlada
    # (todos los clientes se envian a ese/esa lista). Si no viene, cada cliente
    # recibe en sus correos normales del Excel.
    test_to = os.getenv("TEST_EMAIL_TO", "").strip()
    # Ignorar DB_FECHA_PLAN fija del .env: en produccion SIEMPRE se usa "ayer"
    # (a las 3 AM buscamos el plan del dia anterior).
    os.environ.pop("DB_FECHA_PLAN", None)
    if not os.getenv("DB_FECHA_PLAN_AUTO", "").strip():
        os.environ["DB_FECHA_PLAN_AUTO"] = "ayer"
    if test_to:
        log.info(
            "Modo EJECUCION_NOCTURNA: todos los clientes, envío por correo activo, "
            "REDIRIGIDO a TEST_EMAIL_TO=%s, DB_FECHA_PLAN_AUTO=%s, "
            "filtro propietario = nombre del Excel por fila.",
            test_to,
            os.getenv("DB_FECHA_PLAN_AUTO", "ayer"),
        )
    else:
        log.info(
            "Modo EJECUCION_NOCTURNA: todos los clientes, envío por correo activo, "
            "destinatarios normales por cliente, DB_FECHA_PLAN_AUTO=%s, "
            "filtro propietario = nombre del Excel por fila.",
            os.getenv("DB_FECHA_PLAN_AUTO", "ayer"),
        )


def _build_like_pattern_from_cliente(nombre_cliente: str) -> str:
    """
    Patrón ILIKE para propietario.
    Por defecto: %<cliente>% para que coincida “contiene”.
    """
    n = (nombre_cliente or "").strip()
    if not n:
        return "%"
    if "%" in n or "_" in n:
        # Si el cliente ya trae comodines, lo respetamos.
        return n
    return f"%{n}%"


def _normalizar_nombre(texto: str) -> str:
    """Normaliza nombres para comparaciones flexibles (sin tildes/símbolos)."""
    t = unicodedata.normalize("NFKD", str(texto or ""))
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = re.sub(r"[^a-zA-Z0-9]+", " ", t).strip().lower()
    t = re.sub(r"\s+", " ", t)
    return t


def _cliente_coincide_con_activos(nombre_cliente: str, activos_norm: Set[str]) -> bool:
    """True si el nombre del cliente coincide (contains bidireccional) con activos."""
    nc = _normalizar_nombre(nombre_cliente)
    if not nc:
        return False
    for na in activos_norm:
        if nc in na or na in nc:
            return True
    return False


def obtener_propietarios_activos_en_fecha(log: logging.Logger) -> Set[str]:
    """
    Retorna propietarios con sacrificio para FECHA_PLAN (rango [fecha, fecha+2)).
    Se usa como prefiltro para no consultar los 453 clientes uno por uno.
    """
    fecha_plan = resolver_db_fecha_plan()
    if not fecha_plan:
        return set()

    psycopg = _obtener_psycopg()
    dsn_log, kwargs = _conn_postgres_desde_env()
    sql = """
        SELECT DISTINCT v.nombre_propietario
        FROM trazabilidad_proceso.plan_faena pf
        JOIN trazabilidad_proceso.plan_faena_producto pfp
          ON pfp.id_plan_faena = pf.id
        JOIN trazabilidad_proceso.vw_pbi01 v
          ON v.codigo = pfp.id_producto
        WHERE pf.fecha_plan = %s::date
          AND v.fecha_insensibilizacion >= %s::timestamp
          AND v.fecha_insensibilizacion < (%s::date + INTERVAL '2 days')
          AND NULLIF(BTRIM(COALESCE(v.nombre_propietario, '')), '') IS NOT NULL
    """
    try:
        log.debug("Prefiltro activos: conectando a PostgreSQL: %s", dsn_log)
        with psycopg.connect(**kwargs) as conn:
            with conn.cursor() as cur:
                cur.execute(sql, (fecha_plan, fecha_plan, fecha_plan))
                rows = cur.fetchall()
                activos = {str(r[0]).strip() for r in rows if r and r[0] is not None and str(r[0]).strip()}
                return activos
    except Exception as e:
        log.warning("No se pudo calcular prefiltro de propietarios activos; se usará lista completa. Motivo: %s", e)
        return set()


def _render_sql_placeholders(
    sql_raw: str,
    *,
    nombre_cliente: str,
    fecha_plan: Optional[str],
    propietario_like: Optional[str],
) -> Tuple[str, List[Any]]:
    """
    Convierte placeholders {{TOKEN}} a placeholders del driver y arma parámetros.

    Placeholders soportados:
      - {{CLIENTE}}: nombre_cliente
      - {{FECHA_PLAN}}: string 'YYYY-MM-DD' (se castea a date en SQL o se usa como DATE)
      - {{PROPIETARIO_LIKE}}: patrón ILIKE (por defecto %<cliente>%)
      - {{USE_PBI02}}: TRUE/FALSE para habilitar/deshabilitar el JOIN a vw_pbi02
                      (fuente opcional del prefijo XxY). Default: FALSE (vw_pbi02
                      es muy pesada y puede tardar >10 min).

    Nota: el SQL debe usar esos placeholders en vez de literales fijos en el CTE params.
    """
    # Postgres (psycopg) usa %s
    params: List[Any] = []

    def put(val: Any) -> str:
        params.append(val)
        return "%s"

    # psycopg usa %s; cualquier '%' literal en el SQL debe escaparse como '%%'
    # (p.ej. LIKE 'CAVA%'). Escapamos TODO y luego insertamos placeholders %s.
    sql = sql_raw.replace("%", "%%")
    if "{{CLIENTE}}" in sql:
        sql = sql.replace("{{CLIENTE}}", put(nombre_cliente))

    if "{{FECHA_PLAN}}" in sql:
        if not fecha_plan:
            raise ValueError("Falta DB_FECHA_PLAN (YYYY-MM-DD) para {{FECHA_PLAN}}.")
        sql = sql.replace("{{FECHA_PLAN}}", put(fecha_plan))

    if "{{PROPIETARIO_LIKE}}" in sql:
        pat = propietario_like.strip() if propietario_like and propietario_like.strip() else _build_like_pattern_from_cliente(nombre_cliente)
        sql = sql.replace("{{PROPIETARIO_LIKE}}", put(pat))

    if "{{USE_PBI02}}" in sql:
        raw = os.getenv("USE_PBI02", "0").strip().lower()
        use_pbi02 = raw in ("1", "true", "yes", "si", "sí")
        # Interpolamos booleano como literal seguro (no viene de usuario externo).
        sql = sql.replace("{{USE_PBI02}}", "TRUE" if use_pbi02 else "FALSE")

    return sql, params


def _obtener_pyodbc():
    try:
        import pyodbc  # type: ignore

        return pyodbc
    except ImportError as e:
        raise RuntimeError(
            "Falta 'pyodbc'. Instala con: .\\.venv\\Scripts\\python.exe -m pip install pyodbc"
        ) from e


def _obtener_psycopg():
    try:
        import psycopg  # type: ignore

        return psycopg
    except ImportError as e:
        raise RuntimeError(
            "Falta 'psycopg'. Instala con: .\\.venv\\Scripts\\python.exe -m pip install \"psycopg[binary]\""
        ) from e


def _tiene_config_postgres() -> bool:
    return bool(os.getenv("POSTGRES_HOST", "").strip()) and bool(os.getenv("POSTGRES_DB", "").strip())


def _conn_postgres_desde_env() -> Tuple[str, dict]:
    """
    Devuelve (dsn_sin_password_para_log, kwargs_para_psycopg.connect).
    """
    host = os.getenv("POSTGRES_HOST", "").strip()
    db = os.getenv("POSTGRES_DB", "").strip()
    user = os.getenv("POSTGRES_USER", "").strip()
    password = os.getenv("POSTGRES_PASSWORD", "")
    port_raw = os.getenv("POSTGRES_PORT", "5432").strip()
    port = int(port_raw) if port_raw.isdigit() else 5432

    if not host or not db:
        raise ValueError("Faltan POSTGRES_HOST o POSTGRES_DB.")
    if not user or not password:
        raise ValueError("Faltan POSTGRES_USER o POSTGRES_PASSWORD.")

    dsn_log = f"postgresql://{user}:***@{host}:{port}/{db}"
    kwargs = {
        "host": host,
        "port": port,
        "dbname": db,
        "user": user,
        "password": password,
        # Evita problemas frecuentes de SSL en redes internas; puedes ajustar si tu servidor requiere SSL.
        "sslmode": os.getenv("POSTGRES_SSLMODE", "prefer"),
    }
    return dsn_log, kwargs


def ejecutar_consulta_db_por_cliente(
    nombre_cliente: str,
    log: logging.Logger,
) -> Tuple[List[str], List[Tuple[Any, ...]]]:
    """
    Ejecuta la consulta SQL y devuelve (columnas, filas).

    Conexión:
      - DB_CONN_STR: connection string ODBC completo (recomendado)

    Consulta:
      - DB_QUERY o DB_QUERY_FILE

    Parámetro opcional por cliente:
      - Si tu SQL incluye el placeholder {{CLIENTE}}, se reemplaza por '?'
        y se ejecuta como consulta parametrizada pasando nombre_cliente.
    """
    sql_raw = _leer_sql_desde_env_o_archivo()
    db_type = os.getenv("DB_TYPE", "").strip().lower()

    # Preferir PostgreSQL si hay variables POSTGRES_* o DB_TYPE=postgres.
    if db_type in ("postgres", "postgresql") or (_tiene_config_postgres() and not os.getenv("DB_CONN_STR", "").strip()):
        psycopg = _obtener_psycopg()
        dsn_log, kwargs = _conn_postgres_desde_env()
        fecha_plan = resolver_db_fecha_plan()
        propietario_like = os.getenv("DB_PROPIETARIO_LIKE", "").strip() or None
        sql, params = _render_sql_placeholders(
            sql_raw,
            nombre_cliente=nombre_cliente,
            fecha_plan=fecha_plan,
            propietario_like=propietario_like,
        )

        log.debug("Conectando a PostgreSQL: %s", dsn_log)
        # Reintentos cortos para "conflict with recovery" (hot standby / réplica).
        max_intentos = int(os.getenv("DB_MAX_RETRIES", "4").strip() or "4")
        ultimo_error: Optional[Exception] = None
        for intento in range(1, max_intentos + 1):
            try:
                with psycopg.connect(**kwargs) as conn:
                    with conn.cursor() as cur:
                        if params:
                            cur.execute(sql, params)
                        else:
                            cur.execute(sql)
                        cols = [d.name for d in (cur.description or [])]
                        filas = [tuple(r) for r in cur.fetchall()]
                        return cols, filas
            except Exception as e:
                msg = str(e).lower()
                if "conflict with recovery" in msg or "must be removed" in msg:
                    ultimo_error = e
                    espera = 2 ** (intento - 1)
                    log.warning(
                        "PostgreSQL canceló la consulta por recuperación (intento %s/%s). Reintentando en %ss...",
                        intento,
                        max_intentos,
                        espera,
                    )
                    import time

                    time.sleep(espera)
                    continue
                # Propagamos un error claro para cortar rápido en caso de credenciales inválidas.
                if "password authentication failed" in msg:
                    raise RuntimeError(
                        "Conexión a PostgreSQL rechazada: contraseña inválida. "
                        "Revisa POSTGRES_USER/POSTGRES_PASSWORD en tu .env."
                    ) from e
                raise
        # Si salimos del for por reintentos agotados, propagamos un mensaje claro.
        raise RuntimeError(
            f"PostgreSQL canceló la consulta por recuperación (standby) tras {max_intentos} intentos. "
            "La consulta está tardando demasiado y la réplica la cancela. "
            "Soluciones: (1) optimizar el SQL, (2) ejecutar contra el master, o "
            "(3) subir DB_MAX_RETRIES y esperar a que la réplica libere la recuperación."
        ) from ultimo_error

    # Fallback ODBC (SQL Server / etc.)
    conn_str = os.getenv("DB_CONN_STR", "").strip()
    if not conn_str:
        raise ValueError(
            "No hay configuración de BD. Define POSTGRES_* (PostgreSQL) o DB_CONN_STR (ODBC)."
        )
    # Para ODBC mantenemos soporte mínimo solo para {{CLIENTE}}
    usa_param_cliente = "{{CLIENTE}}" in sql_raw
    sql = sql_raw.replace("{{CLIENTE}}", "?") if usa_param_cliente else sql_raw

    pyodbc = _obtener_pyodbc()
    log.debug("Conectando a BD vía ODBC...")
    with pyodbc.connect(conn_str, timeout=60) as conn:
        cur = conn.cursor()
        if usa_param_cliente:
            cur.execute(sql, (nombre_cliente,))
        else:
            cur.execute(sql)

        cols = [d[0] for d in (cur.description or [])]
        rows = cur.fetchall()
        filas = [tuple(r) for r in rows]
        return cols, filas


def aplicar_guardrail_duplicados_por_animal(
    columnas: Sequence[str],
    filas: List[Tuple[Any, ...]],
    nombre_cliente: str,
    log: logging.Logger,
) -> Tuple[List[Tuple[Any, ...]], int]:
    """
    Detecta y elimina filas duplicadas por `Animal` para evitar adjuntos inflados.

    Mantiene la primera ocurrencia por animal (orden estable) y registra en log
    cuántas filas se removieron para trazabilidad.
    """
    if not filas:
        return filas, 0

    idx_animal: Optional[int] = None
    for i, col in enumerate(columnas):
        if str(col).strip().lower() == "animal":
            idx_animal = i
            break
    if idx_animal is None:
        # Sin columna Animal no podemos validar de forma segura.
        log.warning(
            "Guardrail duplicados: no se encontró columna 'Animal' para cliente '%s'; se omite validación.",
            nombre_cliente[:80],
        )
        return filas, 0

    vistos: Set[Any] = set()
    filtradas: List[Tuple[Any, ...]] = []
    duplicadas = 0
    muestra: List[str] = []

    for fila in filas:
        # Fila inconsistente: no romper flujo, conservarla.
        if idx_animal >= len(fila):
            filtradas.append(fila)
            continue
        animal_id = fila[idx_animal]
        if animal_id in vistos:
            duplicadas += 1
            if len(muestra) < 10:
                muestra.append(str(animal_id))
            continue
        vistos.add(animal_id)
        filtradas.append(fila)

    if duplicadas > 0:
        log.warning(
            "Guardrail duplicados: cliente '%s' tenía %s fila(s) duplicada(s) por Animal; "
            "se removieron automáticamente. Muestra: %s",
            nombre_cliente[:80],
            duplicadas,
            ", ".join(muestra) if muestra else "(sin muestra)",
        )
    return filtradas, duplicadas


def exportar_resultado_a_xlsx(
    columnas: List[str],
    filas: List[Tuple[Any, ...]],
    ruta_salida: Path,
    log: logging.Logger,
    nombre_hoja: str = "DATA",
) -> Path:
    """Exporta un resultset tabular a un .xlsx con cabecera. Devuelve la ruta final escrita."""
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    # Si el archivo ya existe y está abierto (bloqueado), intentamos escribir con nombre alterno.
    ruta_final = ruta_salida
    if ruta_final.exists():
        try:
            ruta_final.unlink()
        except PermissionError:
            base = ruta_salida.with_suffix("")
            suf = ruta_salida.suffix
            for i in range(1, 50):
                candidata = Path(f"{base}_{i}{suf}")
                if not candidata.exists():
                    ruta_final = candidata
                    log.warning(
                        "Archivo de salida en uso (bloqueado). Se guardará como: %s",
                        ruta_final,
                    )
                    break
    usar_template = os.getenv("DB_EXPORT_USE_TEMPLATE", "").strip().lower() in ("1", "true", "yes")
    template_path_raw = os.getenv("DB_TEMPLATE_XLSM_PATH", "").strip().strip('"').strip("'")
    if usar_template and template_path_raw:
        template_path = Path(template_path_raw).expanduser().resolve()
        if template_path.is_file():
            return exportar_resultado_a_xlsx_con_template(
                columnas=columnas,
                filas=filas,
                ruta_salida=ruta_final,
                log=log,
                template_path=template_path,
                sheet_name="Rendimientos",
                nombre_hoja=nombre_hoja,
            )

    # Fallback: Excel genérico (estructura mínima)
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja[:31] if nombre_hoja else "DATA"

    if columnas:
        ws.append(columnas)
    for f in filas:
        ws.append(list(f))

    # Auto-ajuste simple de anchos (capado para evitar monstruos)
    max_cols = max(1, len(columnas) if columnas else (len(filas[0]) if filas else 1))
    for c in range(1, max_cols + 1):
        letra = get_column_letter(c)
        max_len = 0
        for cell in ws[letra]:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[letra].width = min(max(10, max_len + 2), 60)

    wb.save(ruta_final)
    log.info("Reporte (DB) guardado: %s", ruta_final)
    return ruta_final


def _normalizar_valor_excel(v: Any) -> Any:
    """Convierte tipos (time/datetime) a algo que Excel/OpenPyXL maneje bien."""
    if v is None:
        return None
    try:
        import datetime as _dt

        if isinstance(v, _dt.datetime):
            v = v.time()
        if isinstance(v, _dt.time):
            return f"{v.hour:02d}:{v.minute:02d}"
    except Exception:
        pass
    return v


def _fmt_fecha_texto(v: Any) -> Optional[str]:
    """Devuelve fecha como texto YYYY-MM-DD (robusto para plantillas con formatos raros)."""
    if v is None:
        return None
    try:
        import datetime as _dt

        if isinstance(v, _dt.datetime):
            v = v.date()
        if isinstance(v, _dt.date):
            return v.strftime("%Y-%m-%d")
    except Exception:
        pass
    s = str(v).strip()
    return s or None


def _fmt_hora_texto(v: Any) -> Optional[str]:
    """Devuelve hora como texto HH:MM (robusto para plantillas con formatos raros)."""
    if v is None:
        return None
    try:
        import datetime as _dt

        if isinstance(v, _dt.datetime):
            v = v.time()
        if isinstance(v, _dt.time):
            # Incluye segundos y milisegundos si vienen
            ms = int(v.microsecond / 1000)
            return f"{v.hour:02d}:{v.minute:02d}:{v.second:02d}.{ms:03d}"
    except Exception:
        pass
    s = str(v).strip()
    # Si viene con segundos/milisegundos, recorta a HH:MM si es posible
    if len(s) >= 8 and s[2] == ":":
        return s
    if len(s) >= 5 and s[2] == ":":
        return s[:5] + ":00.000"
    return s or None


def _to_date(v: Any):
    """Convierte a datetime.date si es posible."""
    if v is None:
        return None
    try:
        import datetime as _dt

        if isinstance(v, _dt.datetime):
            return v.date()
        if isinstance(v, _dt.date):
            return v
        s = str(v).strip()
        # formatos comunes
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return _dt.datetime.strptime(s[:10], fmt).date()
            except Exception:
                pass
    except Exception:
        pass
    return None


def _to_time(v: Any):
    """Convierte a datetime.time si es posible (con microsegundos)."""
    if v is None:
        return None
    try:
        import datetime as _dt

        if isinstance(v, _dt.datetime):
            return v.time()
        if isinstance(v, _dt.time):
            return v
        s = str(v).strip()
        # 14:33:55.078 / 14:33:55 / 14:33
        for fmt in ("%H:%M:%S.%f", "%H:%M:%S", "%H:%M"):
            try:
                return _dt.datetime.strptime(s, fmt).time()
            except Exception:
                pass
    except Exception:
        pass
    return None


def _coerce_float_excel(val: Any) -> float:
    try:
        if val is None:
            return 0.0
        return float(val)
    except Exception:
        return 0.0


def _rellenar_bloque_resumen_plantilla_fr(
    ws: Any,
    columnas: List[str],
    filas: List[Tuple[Any, ...]],
) -> None:
    """
    Actualiza el bloque superior como en `FRF_Rend ACTUALIZADO.xlsm` (Rendimientos):
    C9 nombre propietario, C10 cantidad de animales, C11 fecha, C12 hora (tipo time);
    N3 suma peso pie, N4 suma pesos medias canal, N5 % canal/pie, N6/N7 Macho/Hembra, N8 promedio pie.
    """
    if not filas:
        return
    idx = {c: i for i, c in enumerate(columnas)}
    pi = idx.get("Propietario")
    if pi is not None:
        for f in filas:
            if f[pi] is not None and str(f[pi]).strip():
                ws["C9"].value = str(f[pi]).strip()
                break
    ws["C10"].value = len(filas)
    fi = idx.get("Fecha Sacrificio")
    if fi is None:
        fi = idx.get("Fecha plan")
    if fi is not None:
        d = _to_date(filas[0][fi])
        if d is not None:
            ws["C11"].value = d
            ws["C11"].number_format = "yyyy-mm-dd"
    # Hora de GENERACION del reporte (igual que el Excel oficial: 02:07:33 a.m.
    # cuando la macro lo exporta al cerrar el turno). Antes usabamos la hora del
    # primer sacrificio, pero eso no coincide con el formato oficial.
    import datetime as _dt_now
    ws["C12"].value = _dt_now.datetime.now().time().replace(microsecond=0)
    eidx = idx.get("Peso Animal en Pie (Kg)")
    hidx = idx.get("Peso (Kg.)")
    jidx = idx.get("Peso (Kg.)2")
    s_pie = 0.0
    if eidx is not None:
        s_pie = sum(_coerce_float_excel(f[eidx]) for f in filas)
    s_canal = 0.0
    if hidx is not None:
        s_canal += sum(_coerce_float_excel(f[hidx]) for f in filas)
    if jidx is not None:
        s_canal += sum(_coerce_float_excel(f[jidx]) for f in filas)
    ws["N3"].value = s_pie
    ws["N4"].value = s_canal
    ws["N5"].value = (s_canal / s_pie) if s_pie > 0 else 0.0
    six = idx.get("Sexo")
    macho = 0
    hembra = 0
    if six is not None:
        for f in filas:
            s = str(f[six] or "").lower()
            if "hemb" in s:
                hembra += 1
            else:
                macho += 1
    else:
        macho = len(filas)
    ws["N6"].value = macho
    ws["N7"].value = hembra
    n = len(filas)
    ws["N8"].value = (s_pie / n) if n > 0 else 0.0


def exportar_resultado_a_xlsx_con_template(
    *,
    columnas: List[str],
    filas: List[Tuple[Any, ...]],
    ruta_salida: Path,
    log: logging.Logger,
    template_path: Path,
    sheet_name: str,
    nombre_hoja: str,
) -> Path:
    """
    Exporta el resultset dentro de la misma estructura/formatos de tu plantilla.

    Mapea las columnas de SQL a las columnas fijas de la hoja `Rendimientos`:
      B: Animal, C: Sexo, D: Especie - Raza, E: Peso Animal..., F/G: Fecha/Hora,
      H/I peso/destino 1, J/K destino 2, L/M, N/O, P/Q, R/S, T/U, V rendimiento, W propietario, Y decomiso.
    """
    # Si la plantilla es .xlsm, preserva macros (keep_vba=True) cuando la salida también es .xlsm.
    keep_vba = template_path.suffix.lower() == ".xlsm" and ruta_salida.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(template_path, data_only=False, keep_vba=keep_vba)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{sheet_name}' en la plantilla: {template_path}")
    ws = wb[sheet_name]

    # Logo opcional solo para la hoja de reporte (no para hoja de datos crudos)
    if os.getenv("EXCEL_LOGO_ENABLED", "1").strip().lower() in ("1", "true", "yes"):
        _agregar_logo_a_excel_openpyxl(ws, log)

    header_row = 13
    start_row = header_row + 1

    # Borra valores anteriores en las columnas relevantes (no toques estilos).
    # IMPORTANTE: la plantilla FRF_Rend ACTUALIZADO.xlsm puede traer datos "demo"
    # (p. ej. filas 14..69 con otro cliente). Si la consulta retorna menos filas
    # que la basura pre-cargada, no podemos quedarnos con esas celdas mezclando
    # datos viejos con nuevos. Por eso detectamos DINÁMICAMENTE la última fila
    # con contenido en A/B y limpiamos TODO ese rango.
    cols_relevantes = [
        "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M",
        "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y",
    ]
    # 1) Escaneamos hacia abajo para encontrar la última fila con dato en A o B
    #    (las dos columnas que siempre se llenan en el reporte real).
    #    Limitamos el escaneo a un tope razonable para no leer toda la hoja.
    scan_limit = int(os.getenv("TEMPLATE_CLEAR_SCAN_LIMIT", "2000").strip() or "2000")
    max_scan = min(ws.max_row or start_row, scan_limit)
    ultima_con_dato = start_row - 1
    for r in range(start_row, max_scan + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if (a is not None and a != "") or (b is not None and b != ""):
            ultima_con_dato = r

    # 2) Limpiamos desde start_row hasta el máximo entre:
    #    - la última fila con basura pre-cargada,
    #    - las nuevas filas que vamos a escribir + un colchón de 20,
    #    - un mínimo absoluto (header_row + 50) para que el archivo se vea limpio.
    fin_limpieza = max(
        ultima_con_dato,
        start_row + max(len(filas), 0) + 20,
        start_row + 50,
    )
    log.info(
        "Plantilla: start_row=%s, última fila con dato en A/B detectada=%s, "
        "fin de limpieza=%s, filas a escribir=%s",
        start_row,
        ultima_con_dato,
        fin_limpieza,
        len(filas),
    )
    for r in range(start_row, fin_limpieza + 1):
        for col in cols_relevantes:
            ws[f"{col}{r}"].value = None

    idx = {c: i for i, c in enumerate(columnas)}

    mapping = [
        ("Animal", "B"),
        ("Sexo", "C"),
        ("Especie - Raza", "D"),
        ("Peso Animal en Pie (Kg)", "E"),
        ("Fecha Sacrificio", "F"),
        ("Hora Sacrificio", "G"),
        ("Peso (Kg.)", "H"),
        ("Destino", "I"),
        ("Peso (Kg.)2", "J"),
        ("Destino2", "K"),
        ("Peso (Kg.)3", "L"),
        ("Destino4", "M"),
        ("Peso (Kg.)5", "N"),
        ("Destino6", "O"),
        ("Peso (Kg.)7", "P"),
        ("Destino8", "Q"),
        ("Peso (Kg.)9", "R"),
        ("Destino10", "S"),
        ("Peso (Kg.)12", "T"),
        ("Destino11", "U"),
        ("Rendimiento (%)", "V"),
        ("Propietario", "W"),
        ("Sacrificio de Emergencia", "X"),
        ("Decomiso", "Y"),
    ]

    for i, fila in enumerate(filas):
        r = start_row + i
        # Columna A: # (número correlativo)
        ws[f"A{r}"].value = i + 1
        for alias, col_letter in mapping:
            if alias not in idx:
                continue
            val = fila[idx[alias]]
            if alias == "Fecha Sacrificio":
                d = _to_date(val)
                ws[f"{col_letter}{r}"].value = d if d is not None else _fmt_fecha_texto(val)
                ws[f"{col_letter}{r}"].number_format = "yyyy-mm-dd"
            elif alias == "Hora Sacrificio":
                # Hora como TEXTO para evitar serial decimal por formatos del template
                ws[f"{col_letter}{r}"].value = _fmt_hora_texto(val)
                ws[f"{col_letter}{r}"].number_format = "@"
            else:
                ws[f"{col_letter}{r}"].value = _normalizar_valor_excel(val)

    # Bloque Empresa / Fecha / Totales (M3:N8) como plantilla FRF_Rend ACTUALIZADO.xlsm
    if os.getenv("EXCEL_FILL_TEMPLATE_SUMMARY", "1").strip().lower() in ("1", "true", "yes"):
        try:
            _rellenar_bloque_resumen_plantilla_fr(ws, columnas, filas)
        except Exception as ex:
            log.warning("No se pudo rellenar el resumen superior de la plantilla: %s", ex)

    # Hoja DATA (tabla cruda de la consulta) para que el archivo sea "definitivo":
    # Reporte + Datos en el mismo Excel.
    if os.getenv("INCLUIR_HOJA_DATA", "1").strip().lower() in ("1", "true", "yes"):
        nombre_data = os.getenv("DATA_SHEET_NAME", "DATA").strip() or "DATA"
        try:
            if nombre_data in wb.sheetnames:
                ws_data = wb[nombre_data]
                # Limpia contenido previo (mantiene estilos simples si existieran)
                ws_data.delete_rows(1, ws_data.max_row or 1)
            else:
                ws_data = wb.create_sheet(title=nombre_data[:31])
        except Exception:
            ws_data = wb.create_sheet(title="DATA")

        # Header
        if columnas:
            ws_data.append(list(columnas))
        # Rows
        for f in filas:
            ws_data.append([_normalizar_valor_excel(x) for x in f])

        # Si solo quieres enviar la hoja de datos (sin otras pestañas), puedes controlar esto por env.
        # Ejemplo: EXCEL_KEEP_SHEETS=Datos  (y DATA_SHEET_NAME=Datos)
        keep_raw = os.getenv("EXCEL_KEEP_SHEETS", "").strip()
        if keep_raw:
            keep = {x.strip() for x in keep_raw.split(",") if x.strip()}
            # Normaliza por si el usuario puso "DATA" pero la hoja se llama distinto
            if "DATA" in keep and nombre_data not in keep:
                keep.add(nombre_data)
            for sn in list(wb.sheetnames):
                if sn not in keep:
                    try:
                        del wb[sn]
                    except Exception:
                        pass

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_salida)
    log.info("Reporte (DB) guardado con plantilla: %s", ruta_salida)
    return ruta_salida


def _agregar_logo_a_excel_openpyxl(ws, log: logging.Logger) -> None:
    """
    Inserta un logo en la hoja usando openpyxl.
    Controlado por variables:
      - EXCEL_LOGO_PATH (default: Colbeef.png junto al script)
      - EXCEL_LOGO_CELL (default: B2)
      - EXCEL_LOGO_WIDTH_PX / EXCEL_LOGO_HEIGHT_PX (default: 160x40)
      - EXCEL_LOGO_FORCE (default: 0) -> si ya hay imágenes, no agrega otra
    """
    try:
        from openpyxl.drawing.image import Image as XLImage  # type: ignore
    except Exception:
        # Sin Pillow, openpyxl no puede manejar imágenes.
        return

    force = os.getenv("EXCEL_LOGO_FORCE", "").strip().lower() in ("1", "true", "yes")
    try:
        existentes = getattr(ws, "_images", None)
        if not force and existentes and len(existentes) > 0:
            return
    except Exception:
        pass

    base = directorio_aplicacion()
    logo_raw = os.getenv("EXCEL_LOGO_PATH", str(base / "Colbeef.png")).strip().strip('"').strip("'")
    if not logo_raw:
        return
    logo_path = Path(logo_raw)
    if not logo_path.is_absolute():
        logo_path = (base / logo_path).resolve()
    if not logo_path.is_file():
        log.warning("Logo no encontrado (se omite): %s", logo_path)
        return

    cell = os.getenv("EXCEL_LOGO_CELL", "B2").strip() or "B2"
    try:
        w = int(os.getenv("EXCEL_LOGO_WIDTH_PX", "160").strip() or "160")
        h = int(os.getenv("EXCEL_LOGO_HEIGHT_PX", "40").strip() or "40")
    except Exception:
        w, h = 160, 40

    try:
        img = XLImage(str(logo_path))
        img.width = max(10, w)
        img.height = max(10, h)
        ws.add_image(img, cell)
    except Exception as e:
        log.warning("No se pudo insertar logo en Excel: %s", e)


def cargar_adjunto_binario(ruta: Path) -> Tuple[bytes, str]:
    """Lee el archivo generado en modo binario para adjuntarlo al correo."""
    with open(ruta, "rb") as f:
        datos = f.read()
    return datos, ruta.name


def mime_para_extension(nombre_archivo: str) -> str:
    """Subtipo MIME para el adjunto según la extensión."""
    ext = Path(nombre_archivo).suffix.lower()
    if ext == ".xlsx":
        return "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if ext == ".xlsm":
        return "vnd.ms-excel.sheet.macroEnabled.12"
    if ext == ".pdf":
        return "application/pdf"
    return "octet-stream"


# Alineación centrada (Excel COM / VBA xlCenter)
_XL_ALIGN_CENTER = -4108
# Dirección xlUp para buscar última fila con datos
_XL_UP = -4162
_XL_VALUES = 2
_XL_WHOLE = 1
_XL_PART = 2


def _com_ultima_fila_columna(ws, col: str) -> int:
    """Última fila con valor en una columna (Ctrl+↑ desde el fondo)."""
    try:
        return int(ws.Range(f"{col}{ws.Rows.Count}").End(_XL_UP).Row)
    except Exception:
        return 1


def _com_find_first_row(ws, what: str, *, look_at: int = _XL_PART) -> Optional[int]:
    """Encuentra la primera fila donde aparezca `what` en la hoja (valores visibles)."""
    try:
        found = ws.Cells.Find(
            What=what,
            After=ws.Cells(1, 1),
            LookIn=_XL_VALUES,
            LookAt=look_at,
            SearchOrder=1,  # xlByRows
            SearchDirection=1,  # xlNext
            MatchCase=False,
        )
        if found is None:
            return None
        return int(found.Row)
    except Exception:
        return None


def _com_delete_rows(ws, start_row: int, end_row: int) -> None:
    """Borra un rango de filas (inclusive)."""
    if end_row < start_row:
        return
    try:
        ws.Rows(f"{start_row}:{end_row}").Delete()
    except Exception:
        # Fallback: borrar de a una (más lento pero robusto)
        try:
            for r in range(end_row, start_row - 1, -1):
                ws.Rows(r).Delete()
        except Exception:
            pass


def _com_first_nonempty_row_in_range(ws, *, start_row: int, end_row: int) -> Optional[int]:
    """
    Busca la primera fila con algún valor en A:Y dentro de [start_row, end_row].
    Útil para “pegar” la tabla hacia arriba eliminando huecos.
    """
    if end_row < start_row:
        return None
    try:
        data = ws.Range(f"A{start_row}:Y{end_row}").Value
    except Exception:
        return None
    if data is None:
        return None
    # Normaliza a matriz filas x columnas
    if not isinstance(data, tuple):
        matriz = [list(data) if isinstance(data, (list, tuple)) else [data]]
    elif len(data) > 0 and isinstance(data[0], (tuple, list)):
        matriz = [list(x) for x in data]
    else:
        matriz = [list(data)]
    for i, row in enumerate(matriz):
        if any(_cell_nonempty_pdf(c) for c in row):
            return start_row + i
    return None


def _pdf_padding_filas_print_area() -> int:
    raw = os.getenv("PDF_PRINT_AREA_PADDING_ROWS", "4").strip()
    try:
        return max(0, int(raw))
    except ValueError:
        return 4


def _cell_nonempty_pdf(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, str):
        return bool(v.strip())
    return True


def _com_aplicar_area_impresion_y_formato_tabla(
    ws,
    *,
    header_row: int,
    start_row: int,
    n_filas_datos: int,
    primera_col: str = "A",
    ultima_col: str = "Y",
) -> None:
    """
    Evita huecos enormes en el PDF: no usa UsedRange (a menudo inflado por la plantilla).
    Centra el bloque de datos para que se vea más 'cuadrado'.

    Importante: por defecto NO extiende el área con un "scan" de columnas A/B, porque en
    plantillas viejas suele quedar basura en fila 200+ y el PDF mete cientos de filas vacías
    (hueco gigante entre encabezado y datos).
    """
    if n_filas_datos > 0:
        last_data = start_row + n_filas_datos - 1
    else:
        last_data = start_row
    pad = _pdf_padding_filas_print_area()
    trust = os.getenv("PDF_PRINT_AREA_TRUST_SCAN", "0").strip().lower() in ("1", "true", "yes")
    if trust:
        try:
            last_scan = max(
                _com_ultima_fila_columna(ws, "A"),
                _com_ultima_fila_columna(ws, "B"),
            )
            last_data = max(last_data, last_scan)
        except Exception:
            pass
    print_end = max(header_row, last_data + pad)
    try:
        ws.PageSetup.PrintArea = f"${primera_col}$1:${ultima_col}${print_end}"
    except Exception:
        pass
    try:
        body = ws.Range(f"{primera_col}{start_row}:{ultima_col}{last_data}")
        body.HorizontalAlignment = _XL_ALIGN_CENTER
        body.VerticalAlignment = _XL_ALIGN_CENTER
    except Exception:
        pass
    if os.getenv("PDF_AUTO_FIT_COLUMNS", "").strip().lower() in ("1", "true", "yes"):
        try:
            ws.Range(f"{primera_col}{header_row}:{ultima_col}{print_end}").Columns.AutoFit()
        except Exception:
            pass


def _com_eliminar_saltos_pagina_excel(ws) -> None:
    """ResetAllPageBreaks a veces no basta; borramos uno a uno (H y V)."""
    try:
        ws.ResetAllPageBreaks()
    except Exception:
        pass
    for nombre in ("HPageBreaks", "VPageBreaks"):
        try:
            breaks = getattr(ws, nombre)
            for _ in range(500):
                if breaks.Count <= 0:
                    break
                breaks(1).Delete()
        except Exception:
            pass


def _com_evitar_hueco_vertical_pdf(ws) -> None:
    """
    En PDF suele verse un hueco enorme entre el encabezado verde y los datos cuando:
    - La plantilla dejó PageSetup.CenterVertically = True, o
    - Hay saltos de página manuales heredados.
    """
    try:
        ws.PageSetup.CenterVertically = False
    except Exception:
        pass
    _com_eliminar_saltos_pagina_excel(ws)


def _com_asegurar_filas_datos_visibles(ws, start_row: int, n_filas: int) -> None:
    """La plantilla a veces tiene ocultas filas donde pegamos datos."""
    if n_filas <= 0:
        return
    try:
        last_r = start_row + n_filas - 1
        ws.Rows(f"{start_row}:{last_r}").Hidden = False
    except Exception:
        pass


def _com_ocultar_filas_vacias_detras_tabla(ws, start_row: int, n_filas: int) -> None:
    """
    Filas vacías con altura residual (plantilla / pruebas) inflan el PDF.
    Oculta filas sin contenido en A:Y por debajo del bloque pegado.
    """
    if n_filas <= 0:
        return
    # Por defecto DESACTIVADO: en algunas plantillas ocultar filas rompe la impresión o confunde rangos.
    if os.getenv("PDF_HIDE_BLANK_ROWS_BELOW_TABLE", "0").strip().lower() not in ("1", "true", "yes"):
        return
    r0 = start_row + n_filas
    limite = int(os.getenv("PDF_HIDE_BLANK_ROWS_SCAN_LIMIT", "650").strip() or "650")
    if r0 > limite:
        return
    try:
        data = ws.Range(f"A{r0}:Y{limite}").Value
    except Exception:
        return
    if data is None:
        return
    if not isinstance(data, tuple):
        matriz = [list(data) if isinstance(data, (list, tuple)) else [data]]
    elif len(data) > 0 and isinstance(data[0], (tuple, list)):
        matriz = [list(x) for x in data]
    else:
        matriz = [list(data)]
    for i, row in enumerate(matriz):
        r = r0 + i
        if any(_cell_nonempty_pdf(c) for c in row):
            continue
        try:
            ws.Rows(r).Hidden = True
        except Exception:
            try:
                ws.Rows(r).RowHeight = 1
            except Exception:
                pass


def _com_normalizar_altura_filas_datos(ws, start_row: int, n_filas: int) -> None:
    """Alturas de fila exageradas en la plantilla empujan la tabla hacia abajo en el PDF."""
    if n_filas <= 0:
        return
    if os.getenv("PDF_NORMALIZE_DATA_ROW_HEIGHT", "1").strip().lower() in ("0", "false", "no"):
        return
    last_r = start_row + n_filas - 1
    try:
        pt = float(os.getenv("PDF_DATA_ROW_HEIGHT_PT", "14.25").strip() or "14.25")
        ws.Rows(f"{start_row}:{last_r}").RowHeight = pt
    except Exception:
        pass


def _com_normalizar_altura_filas_rango(ws, start_row: int, end_row: int) -> None:
    """Normaliza altura en un rango (incluye encabezado de tabla)."""
    if end_row < start_row:
        return
    if os.getenv("PDF_NORMALIZE_DATA_ROW_HEIGHT", "1").strip().lower() in ("0", "false", "no"):
        return
    try:
        pt = float(os.getenv("PDF_DATA_ROW_HEIGHT_PT", "14.25").strip() or "14.25")
        ws.Rows(f"{start_row}:{end_row}").RowHeight = pt
    except Exception:
        pass


def _pdf_aplicar_escala_page_setup(ps) -> None:
    """
    Por defecto: encajar ancho a 1 página (tablas anchas).
    PDF_FIT_PAGES_WIDE=0 + PDF_SCALE_PERCENT=80: escala fija (a veces mejora legibilidad).
    """
    fit_wide = os.getenv("PDF_FIT_PAGES_WIDE", "1").strip().lower()
    if fit_wide in ("1", "true", "yes", ""):
        wide_raw = os.getenv("PDF_FIT_PAGES_WIDE_COUNT", "1").strip() or "1"
        try:
            wide_n = max(1, int(wide_raw))
        except ValueError:
            wide_n = 1
        ps.Zoom = False
        ps.FitToPagesWide = wide_n
        return
    scale_raw = os.getenv("PDF_SCALE_PERCENT", "82").strip() or "82"
    try:
        scale = max(10, min(400, int(scale_raw)))
    except ValueError:
        scale = 82
    try:
        ps.Zoom = scale
        ps.FitToPagesWide = False
    except Exception:
        ps.Zoom = False
        ps.FitToPagesWide = 1


def convertir_xlsx_a_pdf_com(
    ruta_xlsx: Path,
    ruta_pdf: Path,
    log: logging.Logger,
) -> Path:
    """
    Convierte un .xlsx a .pdf usando Excel (COM) en Windows.
    Requiere que Excel esté instalado.
    """
    try:
        import win32com.client  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "Para convertir a PDF se requiere 'pywin32' y Excel instalado. "
            "Instala con: .\\.venv\\Scripts\\python.exe -m pip install pywin32"
        ) from e

    ruta_pdf.parent.mkdir(parents=True, exist_ok=True)
    ruta_final = ruta_pdf
    if ruta_final.exists():
        try:
            ruta_final.unlink()
        except PermissionError:
            ruta_final = ruta_pdf.with_name(ruta_pdf.stem + "_1.pdf")

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        excel.AskToUpdateLinks = False
    except Exception:
        pass
    try:
        # 3 = msoAutomationSecurityForceDisable
        excel.AutomationSecurity = 3
    except Exception:
        pass
    try:
        excel.EnableEvents = False
    except Exception:
        pass

    wb = None
    try:
        # Open "defensivo": sin links, sin prompts, y con reparación automática si aplica.
        wb = excel.Workbooks.Open(
            str(ruta_xlsx.resolve()),
            UpdateLinks=0,
            ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
            CorruptLoad=1,
        )
        # Configuración de impresión para un PDF "de reporte" (similar a tu ejemplo).
        sheet_name = os.getenv("PDF_SHEET_NAME", "Rendimientos").strip() or "Rendimientos"
        try:
            ws = wb.Worksheets(sheet_name)
        except Exception:
            ws = wb.Worksheets(1)

        # Constantes Excel
        xlLandscape = 2
        xlPaperA4 = 9
        xlQualityStandard = 0
        xlQualityMinimum = 1
        xlTypePDF = 0

        ps = ws.PageSetup
        ps.Orientation = xlLandscape
        ps.PaperSize = xlPaperA4
        _pdf_aplicar_escala_page_setup(ps)
        # Control de páginas para reducir tamaño:
        # - PDF_FIT_TALL=0/false -> sin límite de páginas hacia abajo (mejor legibilidad)
        # - PDF_FIT_TALL=1       -> fuerza 1 página de alto (muy compacto)
        fit_tall_raw = os.getenv("PDF_FIT_TALL", "0").strip().lower()
        if fit_tall_raw in ("1", "true", "yes"):
            ps.FitToPagesTall = 1
        else:
            ps.FitToPagesTall = False  # sin límite de páginas hacia abajo

        header_last = int(os.getenv("PDF_HEADER_LAST_ROW", "13").strip() or "13")
        start_data = int(os.getenv("PDF_FIRST_DATA_ROW", "14").strip() or "14")
        # Repetir encabezado (logo/títulos + encabezados de tabla)
        ps.PrintTitleRows = f"$1:${header_last}"

        # Márgenes más compactos
        ps.LeftMargin = excel.InchesToPoints(0.25)
        ps.RightMargin = excel.InchesToPoints(0.25)
        ps.TopMargin = excel.InchesToPoints(0.35)
        ps.BottomMargin = excel.InchesToPoints(0.35)
        ps.HeaderMargin = excel.InchesToPoints(0.15)
        ps.FooterMargin = excel.InchesToPoints(0.15)

        ps.CenterHorizontally = True
        ps.CenterVertically = False

        # Pie de página: fecha + paginado
        ps.CenterFooter = "&D  &T"
        ps.RightFooter = "Página &P de &N"

        # Quitar líneas de cuadrícula en impresión
        try:
            ws.PageSetup.PrintGridlines = False
        except Exception:
            pass

        _com_evitar_hueco_vertical_pdf(ws)

        last_a = _com_ultima_fila_columna(ws, "A")
        last_b = _com_ultima_fila_columna(ws, "B")
        last_scan = max(last_a, last_b)
        n_datos = max(0, last_scan - start_data + 1) if last_scan >= start_data else 0
        _com_aplicar_area_impresion_y_formato_tabla(
            ws,
            header_row=header_last,
            start_row=start_data,
            n_filas_datos=n_datos if n_datos > 0 else 1,
        )

        # Exporta esa hoja a PDF (opciones para reducir peso).
        quality_raw = os.getenv("PDF_QUALITY", "minimum").strip().lower()
        quality = xlQualityMinimum if quality_raw in ("min", "minimum", "baja", "low") else xlQualityStandard
        ws.ExportAsFixedFormat(
            xlTypePDF,
            str(ruta_final.resolve()),
            quality,
            False,  # IncludeDocProperties
            False,  # IgnorePrintAreas
        )
        log.info("Reporte PDF guardado: %s", ruta_final)
        return ruta_final
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        finally:
            excel.Quit()


def _encontrar_soffice_exe() -> Optional[Path]:
    """
    Encuentra LibreOffice (soffice) para conversión a PDF sin Excel.
    Devuelve Path al ejecutable o None.
    """
    cand = shutil.which("soffice")
    if cand:
        return Path(cand)
    # Rutas típicas en Windows
    comunes = [
        r"C:\Program Files\LibreOffice\program\soffice.com",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.com",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for p in comunes:
        pp = Path(p)
        if pp.is_file():
            return pp
    return None


def convertir_excel_a_pdf_libreoffice(ruta_excel: Path, ruta_pdf: Path, log: logging.Logger) -> Path:
    """
    Convierte .xlsx/.xlsm a PDF usando LibreOffice (headless).
    No requiere licencia de Excel.
    """
    soffice = _encontrar_soffice_exe()
    if soffice is None:
        raise RuntimeError(
            "No se pudo convertir a PDF sin Excel porque no se encontró LibreOffice. "
            "Instala LibreOffice y vuelve a ejecutar."
        )

    ruta_pdf.parent.mkdir(parents=True, exist_ok=True)
    # LibreOffice siempre escribe en el outdir con mismo nombre base
    outdir = ruta_pdf.parent
    cmd = [
        str(soffice),
        "--headless",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--norestore",
        "--convert-to",
        "pdf",
        "--outdir",
        str(outdir),
        str(ruta_excel),
    ]
    log.info("Convirtiendo a PDF con LibreOffice: %s", " ".join(cmd[:6] + ["..."]))
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
        raise RuntimeError(
            "LibreOffice falló al convertir a PDF. "
            f"returncode={proc.returncode}\nstdout={proc.stdout}\nstderr={proc.stderr}"
        )

    generado = outdir / (ruta_excel.stem + ".pdf")
    if not generado.is_file():
        raise RuntimeError(f"LibreOffice no generó el PDF esperado: {generado}")

    # Renombrar/mover al nombre final solicitado si difiere
    if generado.resolve() != ruta_pdf.resolve():
        try:
            if ruta_pdf.exists():
                ruta_pdf.unlink()
        except Exception:
            pass
        generado.replace(ruta_pdf)
    log.info("Reporte PDF guardado (LibreOffice): %s", ruta_pdf)
    return ruta_pdf


def exportar_pdf_con_template_excel_com(
    *,
    template_xlsm: Path,
    columnas: List[str],
    filas: List[Tuple[Any, ...]],
    ruta_pdf: Path,
    log: logging.Logger,
    sheet_name: str = "Rendimientos",
) -> Path:
    """
    Exporta un PDF directamente desde la plantilla .xlsm usando Excel COM:
    abre plantilla, pega los datos en la tabla y exporta a PDF (sin guardar el xlsm).
    """
    try:
        import win32com.client  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "Para exportar PDF con plantilla se requiere 'pywin32' y Excel instalado. "
            "Instala con: .\\.venv\\Scripts\\python.exe -m pip install pywin32"
        ) from e

    if not template_xlsm.is_file():
        raise FileNotFoundError(f"Plantilla no encontrada: {template_xlsm}")

    ruta_pdf.parent.mkdir(parents=True, exist_ok=True)

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        excel.AskToUpdateLinks = False
    except Exception:
        pass
    try:
        # 3 = msoAutomationSecurityForceDisable
        excel.AutomationSecurity = 3
    except Exception:
        pass
    try:
        excel.EnableEvents = False
    except Exception:
        pass

    wb = None
    try:
        wb = excel.Workbooks.Open(
            str(template_xlsm.resolve()),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
            CorruptLoad=1,
        )

        try:
            ws = wb.Worksheets(sheet_name)
        except Exception:
            ws = wb.Worksheets(1)

        # Si la hoja está protegida, intentar desproteger (sin password).
        try:
            if getattr(ws, "ProtectContents", False):
                ws.Unprotect()
        except Exception:
            pass

        header_row = int(os.getenv("PDF_HEADER_LAST_ROW", "13").strip() or "13")
        start_row = int(os.getenv("PDF_FIRST_DATA_ROW", str(header_row + 1)).strip() or str(header_row + 1))

        # Si la plantilla tiene la tabla MUY abajo (p. ej. fila 240), compactamos el hueco
        # borrando filas vacías entre el encabezado (1..header_row) y la tabla.
        compact = os.getenv("PDF_COMPACT_TABLE_UP", "1").strip().lower() in ("1", "true", "yes")
        if compact:
            # 1) Detecta primer contenido real de la tabla (no solo el texto "Animal")
            # Escaneamos un rango razonable (hasta 400) para no leer toda la hoja.
            scan_end = int(os.getenv("PDF_COMPACT_SCAN_END_ROW", "400").strip() or "400")
            first_content = _com_first_nonempty_row_in_range(
                ws, start_row=header_row + 1, end_row=scan_end
            )
            if first_content is not None and first_content > header_row + 1:
                # Borra el hueco completo entre encabezado y primer contenido real.
                _com_delete_rows(ws, header_row + 1, first_content - 1)
                start_row = header_row + 1

        idx = {c: i for i, c in enumerate(columnas)}
        mapping = [
            ("Animal", "B"),
            ("Sexo", "C"),
            ("Especie - Raza", "D"),
            ("Peso Animal en Pie (Kg)", "E"),
            ("Fecha Sacrificio", "F"),
            ("Hora Sacrificio", "G"),
            ("Peso (Kg.)", "H"),
            ("Destino", "I"),
            ("Peso (Kg.)2", "J"),
            ("Destino2", "K"),
            ("Peso (Kg.)3", "L"),
            ("Destino4", "M"),
            ("Peso (Kg.)5", "N"),
            ("Destino6", "O"),
            ("Peso (Kg.)7", "P"),
            ("Destino8", "Q"),
            ("Peso (Kg.)9", "R"),
            ("Destino10", "S"),
            ("Peso (Kg.)12", "T"),
            ("Destino11", "U"),
            ("Rendimiento (%)", "V"),
            ("Propietario", "W"),
            ("Sacrificio de Emergencia", "X"),
            ("Decomiso", "Y"),
        ]

        # Limpia la tabla (valores) en un rango razonable.
        # (No borra estilos).
        # La plantilla puede traer datos demo pre-cargados; detectamos dinámicamente
        # hasta dónde llega esa basura (última fila con contenido en A o B) y
        # limpiamos hasta allí o hasta donde escribamos (lo que sea mayor).
        try:
            scan_limit = int(os.getenv("TEMPLATE_CLEAR_SCAN_LIMIT", "2000").strip() or "2000")
            scan_end_for_clear = min(int(ws.UsedRange.Rows.Count) + start_row, scan_limit)
            col_a_b = ws.Range(f"A{start_row}:B{scan_end_for_clear}").Value
            ultima_con_dato = start_row - 1
            if col_a_b:
                for i, row_vals in enumerate(col_a_b):
                    a_v = row_vals[0] if row_vals and len(row_vals) > 0 else None
                    b_v = row_vals[1] if row_vals and len(row_vals) > 1 else None
                    if (a_v is not None and a_v != "") or (b_v is not None and b_v != ""):
                        ultima_con_dato = start_row + i
        except Exception:
            ultima_con_dato = start_row + 50
        last_row = max(
            ultima_con_dato,
            start_row + max(len(filas), 0) + 20,
            start_row + 50,
        )
        log.info(
            "Plantilla (COM): start_row=%s, última fila con dato en A/B=%s, "
            "fin de limpieza=%s, filas a escribir=%s",
            start_row, ultima_con_dato, last_row, len(filas),
        )
        ws.Range(f"A{start_row}:Y{last_row}").ClearContents()

        # Construye matriz para pegar por columnas (más rápido que celda a celda).
        # Columna A: consecutivo.
        if filas:
            n = len(filas)
            # Pegamos por cada columna mapeada para mantener simple.
            ws.Range(f"A{start_row}:A{start_row+n-1}").Value = [[i + 1] for i in range(n)]
            for alias, col_letter in mapping:
                if alias not in idx:
                    continue
                rng = ws.Range(f"{col_letter}{start_row}:{col_letter}{start_row+n-1}")
                if alias == "Fecha Sacrificio":
                    import datetime as _dt

                    vals = []
                    for fila in filas:
                        d = _to_date(fila[idx[alias]])
                        vals.append([_dt.datetime(d.year, d.month, d.day) if d else _fmt_fecha_texto(fila[idx[alias]])])
                    rng.Value = vals
                    try:
                        rng.NumberFormat = "yyyy-mm-dd"
                    except Exception:
                        try:
                            rng.NumberFormatLocal = "yyyy-mm-dd"
                        except Exception:
                            pass
                elif alias == "Hora Sacrificio":
                    # Hora como TEXTO para que nunca se vea como decimal (0,60625).
                    try:
                        rng.NumberFormat = "@"
                    except Exception:
                        try:
                            rng.NumberFormatLocal = "@"
                        except Exception:
                            pass

                    vals = [[_fmt_hora_texto(fila[idx[alias]])] for fila in filas]
                    rng.Value = vals
                else:
                    col_values = [[_normalizar_valor_excel(fila[idx[alias]])] for fila in filas]
                    rng.Value = col_values

        # Pegar tabla al encabezado en el PDF (sin hueco enorme por plantilla)
        n_filas_llenas = len(filas) if filas else 0
        _com_asegurar_filas_datos_visibles(ws, start_row, n_filas_llenas)
        _com_ocultar_filas_vacias_detras_tabla(ws, start_row, n_filas_llenas)
        # Normaliza también la fila de encabezado de tabla (start_row) para que no empuje hacia abajo
        end_norm = start_row + max(n_filas_llenas, 1) - 1
        _com_normalizar_altura_filas_rango(ws, start_row, end_norm)
        _com_evitar_hueco_vertical_pdf(ws)

        # Reutiliza la misma configuración de impresión bonita que en convertir_xlsx_a_pdf_com
        os.environ.setdefault("PDF_SHEET_NAME", str(sheet_name))
        # Exporta PDF usando la función existente (pero sobre este workbook/worksheet ya abierto)
        # -> replicamos config aquí para evitar re-abrir.
        xlLandscape = 2
        xlPaperA4 = 9
        xlQualityStandard = 0
        xlQualityMinimum = 1
        xlTypePDF = 0

        ps = ws.PageSetup
        ps.Orientation = xlLandscape
        ps.PaperSize = xlPaperA4
        try:
            ps.CenterVertically = False
        except Exception:
            pass
        _pdf_aplicar_escala_page_setup(ps)
        fit_tall_raw = os.getenv("PDF_FIT_TALL", "0").strip().lower()
        if fit_tall_raw in ("1", "true", "yes"):
            ps.FitToPagesTall = 1
        else:
            ps.FitToPagesTall = False
        ps.PrintTitleRows = f"$1:${header_row}"

        ps.LeftMargin = excel.InchesToPoints(0.25)
        ps.RightMargin = excel.InchesToPoints(0.25)
        ps.TopMargin = excel.InchesToPoints(0.35)
        ps.BottomMargin = excel.InchesToPoints(0.35)
        ps.HeaderMargin = excel.InchesToPoints(0.15)
        ps.FooterMargin = excel.InchesToPoints(0.15)
        ch = os.getenv("PDF_CENTER_PAGE_HORIZ", "1").strip().lower() not in ("0", "false", "no")
        ps.CenterHorizontally = ch
        ps.CenterFooter = "&D  &T"
        ps.RightFooter = "Página &P de &N"
        try:
            ws.PageSetup.PrintGridlines = False
        except Exception:
            pass
        _com_aplicar_area_impresion_y_formato_tabla(
            ws,
            header_row=header_row,
            start_row=start_row,
            n_filas_datos=n_filas_llenas,
        )

        # Si el área de impresión queda mal, Excel puede “cortar” las filas con datos en el PDF.
        # Por defecto ignoramos PrintArea al exportar (sí respeta filas ocultas y márgenes).
        ign_print = os.getenv("PDF_EXPORT_IGNORE_PRINT_AREA", "1").strip().lower() not in (
            "0",
            "false",
            "no",
        )
        last_print = (
            start_row + n_filas_llenas - 1 + _pdf_padding_filas_print_area()
            if n_filas_llenas > 0
            else start_row
        )
        log.info(
            "PDF plantilla: filas SQL pegadas=%s, fila inicio datos=%s, ~última fila contenido=%s, "
            "IgnorePrintAreas=%s",
            n_filas_llenas,
            start_row,
            last_print,
            ign_print,
        )

        quality_raw = os.getenv("PDF_QUALITY", "minimum").strip().lower()
        quality = xlQualityMinimum if quality_raw in ("min", "minimum", "baja", "low") else xlQualityStandard
        ws.ExportAsFixedFormat(
            xlTypePDF,
            str(ruta_pdf.resolve()),
            quality,
            False,  # IncludeDocProperties
            ign_print,  # IgnorePrintAreas (True = no recortar por PrintArea; evita tabla vacía)
        )
        log.info("Reporte PDF (plantilla) guardado: %s", ruta_pdf)
        return ruta_pdf
    finally:
        try:
            if wb is not None:
                try:
                    wb.Saved = True
                except Exception:
                    pass
                wb.Close(SaveChanges=False)
        finally:
            excel.Quit()


def construir_asunto(nombre_cliente: str) -> str:
    """Asunto dinámico por cliente y fecha de faena (si aplica)."""
    pref = os.getenv("EMAIL_SUBJECT_PREFIX", "Rendimientos Beneficio").strip() or "Rendimientos Beneficio"
    fecha_plan = resolver_db_fecha_plan() or date.today().strftime("%Y-%m-%d")
    return f"{pref} - {nombre_cliente} - {fecha_plan}"


def construir_cuerpo_html(nombre_cliente: str, total_canales: int, fecha_beneficio: str) -> str:
    """
    Cuerpo HTML personalizado por cliente.
    Formato alineado al texto solicitado por operación/comercial.

    Si EMAIL_SUBJECT_PREFIX contiene "prueba" o "piloto" y EMAIL_AVISO_PILOTO_CUERPO no es 0,
    se inserta un aviso visible de que el envío es piloto/prueba.
    """
    cliente_esc = html.escape((nombre_cliente or "").strip())
    fecha_esc = html.escape((fecha_beneficio or "").strip())
    canales = max(int(total_canales or 0), 0)
    pref_raw = os.getenv("EMAIL_SUBJECT_PREFIX", "").strip()
    pl = pref_raw.lower()
    banner_piloto = ""
    if os.getenv("EMAIL_AVISO_PILOTO_CUERPO", "1").strip().lower() not in ("0", "false", "no", "off"):
        if "prueba" in pl or "piloto" in pl:
            banner_piloto = """\
              <p style="margin:0 0 20px;padding:12px 14px;background-color:#fff8e6;border:1px solid #e6c200;border-radius:4px;font-size:15px;line-height:1.5;color:#5c4d00;">
                <strong>Aviso:</strong> este env&iacute;o forma parte de una <strong>prueba piloto</strong> del reporte autom&aacute;tico de rendimientos.
                El adjunto corresponde al beneficio indicado; para comentarios pueden escribir a su contacto comercial en Colbeef.
              </p>"""
    return f"""\
<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Reporte automatizado</title>
</head>
<body style="margin:0;padding:0;background-color:#f4f5f7;font-family:Segoe UI, Arial, sans-serif;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background-color:#f4f5f7;padding:24px 16px;">
    <tr>
      <td align="center">
        <table role="presentation" width="600" cellpadding="0" cellspacing="0" style="max-width:600px;background-color:#ffffff;border:1px solid #e2e4e8;border-radius:6px;overflow:hidden;">
          <tr>
            <td style="padding:28px 24px;color:#2d3748;font-size:15px;line-height:1.6;">
              {banner_piloto}
              <p style="margin:0 0 8px;font-size:18px;">Estimados Sres. {cliente_esc}</p>
              <p style="margin:0 0 24px;font-size:18px;">
                Nos complace realizar el env&iacute;o de los Rendimientos de las {canales} Canales de su interes del Beneficio de {fecha_esc}
              </p>
              <p style="margin:0 0 24px;font-size:18px;">Cordialmente..</p>
              <p style="margin:0;font-size:18px;">Sergio Romero</p>
              <p style="margin:0;font-size:18px;">Jefe de Planta Colbeef</p>
              <p style="margin:0;font-size:18px;">Cel. 3154162195</p>
              <p style="margin:0;font-size:18px;">Colbeef S.A.S</p>
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>
"""


def obtener_config_smtp() -> dict:
    """Lee configuración SMTP desde variables de entorno."""
    return {
        "host": os.getenv("SMTP_HOST", ""),
        "port": int(os.getenv("SMTP_PORT", "587")),
        "user": os.getenv("SMTP_USER", ""),
        "password": os.getenv("SMTP_PASSWORD", ""),
        "use_tls": os.getenv("SMTP_USE_TLS", "true").lower() in ("1", "true", "yes"),
        "from_addr": os.getenv("SMTP_FROM", os.getenv("SMTP_USER", "")),
        "from_name": os.getenv("SMTP_FROM_NAME", "").strip(),
    }


def validar_config_smtp(cfg: dict) -> None:
    """Valida que existan los datos mínimos para enviar."""
    if not cfg["host"]:
        raise ValueError("Falta SMTP_HOST en el entorno.")
    if not cfg["user"] or not cfg["password"]:
        raise ValueError("Faltan SMTP_USER o SMTP_PASSWORD en el entorno.")
    if not cfg["from_addr"]:
        raise ValueError("Falta SMTP_FROM o SMTP_USER para el remitente.")


def construir_mensaje(
    de: str,
    para: Sequence[str],
    asunto: str,
    html: str,
    datos_adjunto: bytes,
    nombre_adjunto: str,
    bcc: Optional[Sequence[str]] = None,
) -> MIMEMultipart:
    """Arma el mensaje MIME con HTML y adjunto (.xlsx o .xlsm)."""
    msg = MIMEMultipart()
    msg["Subject"] = asunto
    msg["From"] = de
    msg["To"] = ", ".join(para)
    if bcc:
        msg["Bcc"] = ", ".join(bcc)
    msg.attach(MIMEText(html, "html", "utf-8"))

    sub = mime_para_extension(nombre_adjunto)
    part = MIMEApplication(datos_adjunto, _subtype=sub)
    part.add_header("Content-Disposition", "attachment", filename=nombre_adjunto)
    msg.attach(part)
    return msg


def construir_mensaje_html_sin_adjunto(
    de: str,
    para: Sequence[str],
    asunto: str,
    html: str,
    bcc: Optional[Sequence[str]] = None,
) -> MIMEMultipart:
    """Mensaje solo HTML (p. ej. resumen de ejecución sin adjunto)."""
    msg = MIMEMultipart()
    msg["Subject"] = asunto
    msg["From"] = de
    msg["To"] = ", ".join(para)
    if bcc:
        msg["Bcc"] = ", ".join(bcc)
    msg.attach(MIMEText(html, "html", "utf-8"))
    return msg


def _correos_resumen_ejecucion() -> List[str]:
    """Destinatarios del correo de cierre; vacío = no enviar."""
    raw = os.getenv("REPORT_SUMMARY_EMAIL", "").strip()
    if raw.lower() in ("0", "false", "no", "off"):
        return []
    if raw:
        return _recopilar_correos_unicos(parsear_lista_correos(raw))
    return _recopilar_correos_unicos(
        [
            "desarrollo.tecnologia@colbeef.com",
            "coordinacion.linea@colbeef.com",
        ]
    )


def construir_html_resumen_ejecucion(
    *,
    enviados_ok: int,
    fallidos: int,
    fallos: Sequence[str],
    procesados: int,
    lista_origen: str,
    fecha_beneficio: str,
    dry_run: bool,
) -> str:
    """HTML breve con resultado de la corrida."""
    ok_esc = html.escape(str(enviados_ok))
    bad_esc = html.escape(str(fallidos))
    proc_esc = html.escape(str(procesados))
    orig_esc = html.escape(lista_origen or "")
    fecha_esc = html.escape(fecha_beneficio or "")
    modo = "DRY_RUN (sin envío SMTP real)" if dry_run else "Envío SMTP real"
    modo_esc = html.escape(modo)
    if fallos:
        items = "".join(
            f"<li>{html.escape(n[:200])}</li>" for n in fallos[:80]
        )
        bloque_fallos = f"<p><strong>Clientes con fallo al enviar ({len(fallos)}):</strong></p><ul>{items}</ul>"
        if len(fallos) > 80:
            bloque_fallos += f"<p>… y {len(fallos) - 80} más (ver bot_reportes.log).</p>"
    else:
        bloque_fallos = "<p><strong>Sin fallos registrados en envío.</strong></p>"
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return f"""\
<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8"></head>
<body style="font-family:Segoe UI,Arial,sans-serif;font-size:15px;color:#2d3748;line-height:1.5;">
  <h2 style="margin:0 0 12px;">Resumen — Bot rendimientos</h2>
  <p style="margin:0 0 8px;"><strong>Fecha/hora fin:</strong> {html.escape(ahora)}</p>
  <p style="margin:0 0 8px;"><strong>Modo:</strong> {modo_esc}</p>
  <p style="margin:0 0 8px;"><strong>Fecha beneficio (consulta):</strong> {fecha_esc}</p>
  <p style="margin:0 0 8px;"><strong>Lista clientes:</strong> {orig_esc}</p>
  <p style="margin:0 0 8px;"><strong>Clientes procesados (intentos):</strong> {proc_esc}</p>
  <p style="margin:0 0 8px;"><strong>Envíos correctos:</strong> {ok_esc}</p>
  <p style="margin:0 0 16px;"><strong>Envíos fallidos:</strong> {bad_esc}</p>
  {bloque_fallos}
  <p style="margin-top:20px;font-size:13px;color:#718096;">Log detallado: bot_reportes.log en la carpeta del script.</p>
</body></html>"""


def enviar_correo_resumen_si_aplica(
    cfg: dict,
    log: logging.Logger,
    *,
    enviados_ok: int,
    fallidos: int,
    fallos: Sequence[str],
    procesados: int,
    lista_origen: str,
    dry_run: bool,
) -> None:
    """Al terminar la corrida, notifica por correo el resultado (si hay destinatarios configurados)."""
    destinos = _correos_resumen_ejecucion()
    if not destinos:
        log.debug("Resumen de ejecución: REPORT_SUMMARY_EMAIL desactivado o vacío; no se envía correo de cierre.")
        return
    if procesados <= 0:
        log.debug("Resumen de ejecución: no hubo clientes procesados; no se envía correo de cierre.")
        return
    if dry_run:
        log.info(
            "Resumen de ejecución (no enviado por DRY_RUN): iría a %s — OK=%s fallos=%s",
            destinos,
            enviados_ok,
            fallidos,
        )
        return
    fecha_ben = resolver_db_fecha_plan() or date.today().strftime("%Y-%m-%d")
    asunto = (
        f"Resumen bot rendimientos — {fecha_ben} — "
        f"OK {enviados_ok} / Fallos {fallidos}"
    )
    html = construir_html_resumen_ejecucion(
        enviados_ok=enviados_ok,
        fallidos=fallidos,
        fallos=fallos,
        procesados=procesados,
        lista_origen=lista_origen,
        fecha_beneficio=fecha_ben,
        dry_run=dry_run,
    )
    de = (
        formataddr((cfg.get("from_name", ""), cfg["from_addr"]))
        if cfg.get("from_name")
        else cfg["from_addr"]
    )
    mensaje = construir_mensaje_html_sin_adjunto(
        de=de,
        para=destinos,
        asunto=asunto,
        html=html,
        bcc=None,
    )
    try:
        enviar_correo_smtp(cfg, list(destinos), mensaje)
        log.info("Correo de resumen de ejecución enviado a: %s", ", ".join(destinos))
    except Exception as e:
        log.warning("No se pudo enviar el correo de resumen de ejecución: %s", e)


def enviar_correo_smtp(
    cfg: dict,
    destinatarios_todos: List[str],
    mensaje: MIMEMultipart,
) -> None:
    """Envía el correo; destinatarios_todos incluye To y Bcc para send_message."""
    host = cfg["host"]
    port = cfg["port"]
    user = cfg["user"]
    password = cfg["password"]
    use_tls = cfg["use_tls"]

    if port == 465:
        import ssl

        contexto = ssl.create_default_context()
        with smtplib.SMTP_SSL(host, port, context=contexto) as servidor:
            servidor.login(user, password)
            servidor.send_message(mensaje, to_addrs=destinatarios_todos)
    else:
        with smtplib.SMTP(host, port, timeout=120) as servidor:
            servidor.ehlo()
            if use_tls:
                servidor.starttls()
                servidor.ehlo()
            servidor.login(user, password)
            servidor.send_message(mensaje, to_addrs=destinatarios_todos)


def main() -> int:
    """Punto de entrada: clientes → generación (opcional) → envío por correo."""
    cargar_dotenv_proyecto()
    directorio_script = directorio_aplicacion()
    ruta_log = directorio_script / "bot_reportes.log"
    configurar_logging(ruta_log)

    log = logging.getLogger(__name__)
    log.info("=== Inicio de ejecución del bot de reportes (por cliente) ===")

    aplicar_modo_ejecucion_nocturna(log)

    demo = os.getenv("DEMO_XLSX", "").strip().lower() in ("1", "true", "yes")
    usar_db = os.getenv("USE_DB_QUERY", "").strip().lower() in ("1", "true", "yes")
    dry_run = os.getenv("DRY_RUN", "").strip() in ("1", "true", "yes")
    skip_gen = os.getenv("SKIP_GENERATION", "").strip() in ("1", "true", "yes")
    convertir_pdf = os.getenv("CONVERT_TO_PDF", "").strip().lower() in ("1", "true", "yes")
    borrar_xlsx = os.getenv("DELETE_XLSX_AFTER_PDF", "").strip().lower() in ("1", "true", "yes")
    max_envios_env = os.getenv("MAX_ENVIOS", "").strip()
    max_envios: Optional[int] = int(max_envios_env) if max_envios_env.isdigit() else None

    if demo:
        try:
            dir_salida = obtener_directorio_salida()
            nombre = sanitizar_nombre_archivo(f"DEMO_rendimientos_{date.today().strftime('%Y-%m-%d')}.xlsx")
            ruta_demo = dir_salida / nombre
            generar_excel_demo(ruta_demo, log)
            if convertir_pdf:
                ruta_pdf = ruta_demo.with_suffix(".pdf")
                convertir_xlsx_a_pdf_com(ruta_demo, ruta_pdf, log)
                if borrar_xlsx and ruta_demo.is_file():
                    try:
                        ruta_demo.unlink()
                        log.info("XLSX DEMO eliminado tras PDF: %s", ruta_demo)
                    except PermissionError:
                        log.warning(
                            "No se pudo eliminar el XLSX DEMO (está abierto/bloqueado): %s",
                            ruta_demo,
                        )
            log.info("=== Fin de ejecución (DEMO) ===")
            return 0
        except Exception as exc:
            log.exception("Error en modo DEMO_XLSX: %s: %s", type(exc).__name__, exc)
            log.info("=== Fin de ejecución (con error) ===")
            return 1

    if dry_run:
        log.info("Modo DRY_RUN activo: no se enviarán correos.")

    enviados = 0
    fallidos = 0
    fallos: List[str] = []
    procesados = 0
    try:
        ruta_clientes = obtener_ruta_clientes()
        log.info("Ruta base de clientes: %s", ruta_clientes)

        raw_origen = os.getenv("CLIENTES_LISTA_ORIGEN", "excel").strip().lower()
        if raw_origen in ("excel", "sirt", "merge"):
            lista_origen = raw_origen
        else:
            if raw_origen:
                log.warning(
                    "CLIENTES_LISTA_ORIGEN=%r no es excel|sirt|merge; se usa excel.",
                    raw_origen,
                )
            lista_origen = "excel"
        if lista_origen in ("sirt", "merge"):
            if not usar_db:
                log.warning(
                    "CLIENTES_LISTA_ORIGEN=%s requiere USE_DB_QUERY=1; se usa excel.",
                    lista_origen,
                )
                lista_origen = "excel"
            elif not _tiene_config_postgres():
                log.warning(
                    "CLIENTES_LISTA_ORIGEN=%s requiere POSTGRES_*; se usa excel.",
                    lista_origen,
                )
                lista_origen = "excel"

        if lista_origen == "sirt":
            clientes = cargar_clientes_desde_sirt(log)
        elif lista_origen == "merge":
            excel_clientes = cargar_clientes_desde_excel(ruta_clientes, log)
            sirt_clientes = cargar_clientes_desde_sirt(log)
            clientes = combinar_correos_cliente_sirt_con_excel(
                sirt_clientes, excel_clientes, log
            )
        else:
            clientes = cargar_clientes_desde_excel(ruta_clientes, log)

        if lista_origen in ("sirt", "merge"):
            log.info(
                "Origen de lista de envío: %s (nombres/correos desde SIRT; merge añade correos del Excel si hay match).",
                lista_origen,
            )

        if not clientes:
            raise ValueError("No hay clientes con correo para procesar.")

        solo_cliente = os.getenv("SOLO_CLIENTE_NOMBRE", "").strip()
        if solo_cliente:
            objetivo = _normalizar_nombre_cliente(solo_cliente)
            antes = len(clientes)
            clientes = [
                c for c in clientes if objetivo in _normalizar_nombre_cliente(c.nombre)
            ]
            log.info(
                "Filtro SOLO_CLIENTE_NOMBRE activo: '%s' -> %s/%s filas",
                solo_cliente,
                len(clientes),
                antes,
            )
            if not clientes:
                raise ValueError(
                    f"SOLO_CLIENTE_NOMBRE no encontró coincidencias en la lista de clientes: {solo_cliente!r}"
                )
        elif usar_db and lista_origen == "excel":
            # Prefiltro de rendimiento: procesar solo clientes potencialmente activos
            # en la fecha de faena (evita recorrer los 453 cuando solo unos pocos
            # tuvieron sacrificio/canales ese día).
            activos = obtener_propietarios_activos_en_fecha(log)
            if activos:
                activos_norm = {_normalizar_nombre(n) for n in activos if _normalizar_nombre(n)}
                antes = len(clientes)
                clientes = [c for c in clientes if _cliente_coincide_con_activos(c.nombre, activos_norm)]
                log.info(
                    "Prefiltro activos por fecha aplicado: %s/%s clientes candidatos con sacrificio.",
                    len(clientes),
                    antes,
                )
            else:
                log.info("Prefiltro activos no disponible; se procesará la lista completa de clientes.")

        plantilla = None
        if not usar_db:
            plantilla = obtener_ruta_plantilla_xlsm()
            log.info("Plantilla de macros (.xlsm): %s", plantilla)
        else:
            log.info("Modo BD activo (USE_DB_QUERY=1): se generará .xlsx desde consulta SQL.")

        dir_salida = obtener_directorio_salida()
        log.info("Directorio de salida de reportes: %s", dir_salida)

        cfg = obtener_config_smtp()
        if not dry_run:
            validar_config_smtp(cfg)

        hoy = date.today()

        for cliente in clientes:
            # MAX_ENVIOS limita clientes procesados (éxito o error), no solo "enviados"
            if max_envios is not None and procesados >= max_envios:
                log.info("Límite MAX_ENVIOS (%s) alcanzado; se detiene el proceso.", max_envios)
                break
            procesados += 1

            nombre_adj = nombre_archivo_reporte(cliente.nombre, hoy)
            nombre_adj_seguro = sanitizar_nombre_archivo(nombre_adj)
            ruta_reporte = dir_salida / nombre_adj_seguro
            total_canales_cliente = 0
            fecha_beneficio_correo = resolver_db_fecha_plan() or hoy.strftime("%Y-%m-%d")

            try:
                if skip_gen:
                    log.warning(
                        "SKIP_GENERATION: no se genera Excel para '%s' (omisión intencional).",
                        cliente.nombre[:60],
                    )
                    if not ruta_reporte.is_file():
                        raise FileNotFoundError(
                            "SKIP_GENERATION activo y no existe archivo previo: "
                            f"{ruta_reporte}"
                        )
                elif usar_db:
                    columnas, filas = ejecutar_consulta_db_por_cliente(cliente.nombre, log)
                    filas, duplicadas_removidas = aplicar_guardrail_duplicados_por_animal(
                        columnas,
                        filas,
                        cliente.nombre,
                        log,
                    )
                    total_canales_cliente = len(filas)
                    log.info(
                        "SQL: cliente='%s' -> filas=%d, columnas=%d, duplicados_removidos=%d",
                        cliente.nombre[:60],
                        len(filas),
                        len(columnas),
                        duplicadas_removidas,
                    )
                    if (
                        cliente.cavas_resumen
                        and os.getenv("CLIENTES_LOG_CAVAS", "").strip().lower()
                        in ("1", "true", "yes")
                    ):
                        log.info(
                            "Cavas detectadas (lista SIRT) para '%s': %s",
                            cliente.nombre[:80],
                            cliente.cavas_resumen[:800],
                        )
                    if len(filas) == 0:
                        log.warning(
                            "La consulta no retornó filas para '%s' (FECHA_PLAN=%s, PROPIETARIO_LIKE=%s). "
                            "Se omite generación y envío para este cliente (0 canales).",
                            cliente.nombre[:60],
                            os.getenv("DB_FECHA_PLAN", "(auto)"),
                            os.getenv("DB_PROPIETARIO_LIKE", "(derivado del nombre)"),
                        )
                        continue
                    # Si vamos a PDF y tenemos plantilla .xlsm, generamos el PDF directamente desde la plantilla
                    # con Excel COM (abre la plantilla nativa, rellena la tabla y exporta a PDF sin intermedios).
                    usar_template = os.getenv("DB_EXPORT_USE_TEMPLATE", "").strip().lower() in ("1", "true", "yes")
                    template_path_raw = os.getenv("DB_TEMPLATE_XLSM_PATH", "").strip().strip('"').strip("'")
                    pdf_sin_excel = os.getenv("PDF_SIN_EXCEL", "").strip().lower() in ("1", "true", "yes")
                    if convertir_pdf and usar_template and template_path_raw and not pdf_sin_excel:
                        template_path = Path(template_path_raw).expanduser().resolve()
                        ruta_pdf = ruta_reporte.with_suffix(".pdf")
                        ruta_reporte = exportar_pdf_con_template_excel_com(
                            template_xlsm=template_path,
                            columnas=columnas,
                            filas=filas,
                            ruta_pdf=ruta_pdf,
                            log=log,
                            sheet_name=os.getenv("PDF_SHEET_NAME", "Rendimientos").strip() or "Rendimientos",
                        )
                    else:
                        # Camino sin Excel: llenar plantilla con openpyxl y convertir con LibreOffice (si convertir_pdf=1)
                        # Usamos .xlsm como intermedio si hay plantilla xlsm, para mantener estructura.
                        if convertir_pdf and usar_template and template_path_raw:
                            ruta_xlsx_target = ruta_reporte.with_suffix(".xlsm")
                        else:
                            ruta_xlsx_target = ruta_reporte.with_suffix(".xlsx") if convertir_pdf else ruta_reporte
                        ruta_xlsx = exportar_resultado_a_xlsx(
                            columnas, filas, ruta_xlsx_target, log, nombre_hoja=cliente.nombre
                        )
                        ruta_reporte = ruta_xlsx
                        if convertir_pdf:
                            ruta_pdf = ruta_xlsx.with_suffix(".pdf")
                            if pdf_sin_excel:
                                ruta_reporte = convertir_excel_a_pdf_libreoffice(ruta_xlsx, ruta_pdf, log)
                            else:
                                ruta_reporte = convertir_xlsx_a_pdf_com(ruta_xlsx, ruta_pdf, log)
                            if borrar_xlsx and ruta_xlsx.is_file():
                                try:
                                    ruta_xlsx.unlink()
                                    log.info("XLSX eliminado tras PDF: %s", ruta_xlsx)
                                except PermissionError:
                                    log.warning(
                                        "No se pudo eliminar el XLSX (está abierto/bloqueado): %s",
                                        ruta_xlsx,
                                    )
                else:
                    if plantilla is None:
                        raise RuntimeError("Plantilla no resuelta (estado inválido).")
                    generar_reporte_con_excel_com(
                        plantilla, cliente.nombre, ruta_reporte, log
                    )
            except Exception as gen_err:
                log.exception(
                    "Error al generar reporte para cliente '%s': %s",
                    cliente.nombre,
                    gen_err,
                )
                # Si falla la conexión a BD, no tiene sentido continuar con 400+ clientes.
                if usar_db and isinstance(gen_err, RuntimeError) and "postgresql" in str(gen_err).lower():
                    log.error("Se detiene la ejecución por error de conexión a BD.")
                    break
                continue

            if not ruta_reporte.is_file():
                log.error("No se encontró el archivo generado: %s", ruta_reporte)
                continue

            respaldar_reporte_en_unidad_sincronizada(
                ruta_reporte,
                fecha_subcarpeta=str(fecha_beneficio_correo)[:10],
                log=log,
            )

            datos, nombre_final = cargar_adjunto_binario(ruta_reporte)
            asunto = construir_asunto(cliente.nombre)
            html = construir_cuerpo_html(
                cliente.nombre,
                total_canales_cliente,
                fecha_beneficio_correo,
            )

            # Modo prueba: enviar a un correo fijo sin tocar la base de clientes
            test_to = _parse_env_emails("TEST_EMAIL_TO")
            if test_to:
                destinatarios = test_to
                asunto = os.getenv("TEST_EMAIL_SUBJECT", "").strip() or asunto
                log.info("Modo TEST_EMAIL_TO activo: para=%s (cliente original=%s)", destinatarios, cliente.nombre)
            else:
                destinatarios = list(cliente.correos)
            dest_lower = {e.strip().lower() for e in destinatarios if e and "@" in e}
            bcc_list = [e for e in COPIA_OCULTA if e]
            for em in _parse_env_emails("SMTP_BCC"):
                if em.strip().lower() not in dest_lower:
                    bcc_list.append(em)
            bcc_list = _recopilar_correos_unicos(bcc_list)
            envio_todos = destinatarios + bcc_list

            if dry_run:
                log.info(
                    "[DRY_RUN] No enviado — cliente=%s, adjunto=%s, para=%s",
                    cliente.nombre,
                    nombre_final,
                    destinatarios,
                )
                enviados += 1
                continue

            de = (
                formataddr((cfg.get("from_name", ""), cfg["from_addr"]))
                if cfg.get("from_name")
                else cfg["from_addr"]
            )
            mensaje = construir_mensaje(
                de=de,
                para=destinatarios,
                asunto=asunto,
                html=html,
                datos_adjunto=datos,
                nombre_adjunto=nombre_final,
                bcc=bcc_list if bcc_list else None,
            )
            try:
                enviar_correo_smtp(cfg, envio_todos, mensaje)
                log.info(
                    "Correo enviado a %s (%s destinatario(s) en Para, %s Bcc)",
                    cliente.nombre,
                    len(destinatarios),
                    len(bcc_list),
                )
                enviados += 1
            except Exception as send_err:
                fallidos += 1
                fallos.append(cliente.nombre)
                log.exception("Fallo enviando correo a '%s': %s", cliente.nombre, send_err)

        log.info(
            "Proceso finalizado. Envíos completados (o simulados en DRY_RUN): %s",
            enviados,
        )
        if not dry_run:
            total = enviados + fallidos
            log.info("Resumen envío: OK=%s, FALLIDOS=%s, TOTAL=%s", enviados, fallidos, total)
            if fallos:
                log.warning("Clientes fallidos (%s): %s", len(fallos), ", ".join(fallos[:50]))
        enviar_correo_resumen_si_aplica(
            cfg,
            log,
            enviados_ok=enviados,
            fallidos=fallidos,
            fallos=fallos,
            procesados=procesados,
            lista_origen=lista_origen,
            dry_run=dry_run,
        )
        try:
            marca = directorio_aplicacion() / "ultima_corrida.txt"
            marca.write_text(
                "Última corrida completada (éxito)\n"
                f"hora_fin={datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"enviados_ok={enviados}\n"
                f"fallidos={fallidos}\n"
                f"procesados={procesados}\n"
                f"lista_origen={lista_origen}\n"
                f"dry_run={'1' if dry_run else '0'}\n",
                encoding="utf-8",
            )
        except OSError:
            log.debug("No se pudo escribir ultima_corrida.txt", exc_info=True)
        log.info("=== Fin de ejecución (éxito) ===")
        return 0

    except Exception as exc:
        log.exception(
            "Error global en el bot de reportes: %s: %s",
            type(exc).__name__,
            exc,
        )
        log.info("=== Fin de ejecución (con error) ===")
        return 1


if __name__ == "__main__":
    sys.exit(main())
